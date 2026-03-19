from audit.models import AuditLog, AuditEntityType, AuditAction
from django.core.exceptions import ValidationError
from django.utils.dateparse import parse_date
from django.shortcuts import get_object_or_404
from django.db.models import Avg, Q
from django.utils import timezone
from django.http import HttpResponse
from datetime import datetime

from rest_framework import viewsets, generics, status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView
from drf_spectacular.utils import extend_schema

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from .models.lesson import OnlineLesson
from .models.group import Group
from .models import (
    Student, StudentGroup, StudentPricing, StudentBalances,
    StudentTarnsactions, LeaveReason, StudentGroupLeaves,
    StudentFreezes, StudentBalanceHistory, Attendence,
    Room, Course, Group, GroupTeacher,
    LessonTime, LessonSchedule, Exams, ExamResults,
    TeacherSalaryRules, TeacherSalaryPayments, TeacherSalaryCalculations
)
from .serializers import (
    # Online lesson
    OnlineLessonListSerializer, OnlineLessonDetailSerializer,
    OnlineLessonCreateUpdateSerializer,
    # Student
    AddStudentToGroupSerializer, ActivateStudentSerializer, StudentSearchSerializer,
    StudentSerializer, StudentGroupSerializer, StudentPricingSerializer,
    StudentBalancesSerializer, StudentTarnsactionsSerializer, LeaveReasonSerializer,
    StudentGroupLeavesSerializer, StudentFreezesSerializer, StudentBalanceHistorySerializer,
    AttendenceSerializer, CreateStudentDiscountSerializer, StudentDiscountSerializer,
    # Group
    RoomSerializer, CourseSerializer, GroupSerializer, GroupTeacherSerializer,
    GroupCreateUpdateSerializer,
    # Lesson
    LessonTimeSerializer, LessonScheduleSerializer,
    LessonScheduleDetailSerializer, LessonScheduleListSerializer,
    MoveLessonDateSerializer,
    # Exam
    ExamsSerializer, ExamResultsSerializer,
    ExamCreateSerializer, ExamUpdateSerializer, ExamResultSerializer,
    ExamListSerializer, ExamDetailSerializer, ExamParticipantsSettingsSerializer,
    # Teacher
    TeacherSalaryRulesSerializer, TeacherSalaryPaymentsSerializer,
    TeacherSalaryCalculationsSerializer,
    # Reports
    StudentLeaveReportSerializer, LeaveReportFilterSerializer,
    RestoreLessonSerializer, CancelLessonSerializer,
    DiscountPriceSerializer, BulkLessonTopicSerializer,
    GroupLessonsCalendarSerializer, LessonTopicSerializer,
    # Export
    GroupExportFilterSerializer, GroupExportColumnsSerializer,
    StudentExportFilterSerializer, StudentExportColumnsSerializer,
    GroupExportDataSerializer, StudentExportDataSerializer,
    ExportHistorySerializer, create_export_audit_log,
)
from accounts.models import Employee


# ─── AuditLog yozish uchun helper ────────────────────────────────────────────

def _log(entity_type, entity_id, action, old_data, new_data, user):
    """
    entity_type : 'group' | 'student' | 'payment' | 'lead' | 'user' | 'other'
    action      : 'create' | 'update' | 'delete'
    """
    try:
        employee = user.employee if hasattr(user, 'employee') else None
        AuditLog.objects.create(
            entity_type=entity_type,
            entity_id=entity_id,
            action=action,
            old_data=old_data,
            new_action=new_data,
            performed_by=employee,
            performed_by_role=employee.position if employee else '',
        )
    except Exception:
        pass  # Audit log xatosi asosiy ishni to'xtatmasin


# ════════════════════════════════════════════════════════════════
#  VIEWSETS
# ════════════════════════════════════════════════════════════════

# ── Student ──────────────────────────────────────────────────────

@extend_schema(tags=["Students - Studentlar"])
class StudentViewSet(viewsets.ModelViewSet):
    queryset         = Student.objects.all()
    serializer_class = StudentSerializer

@extend_schema(tags=["StudentGroup - Student va Guruh bog'langan jadval"])
class StudentGroupViewSet(viewsets.ModelViewSet):
    queryset         = StudentGroup.objects.all()
    serializer_class = StudentGroupSerializer

@extend_schema(tags=["StudentPricing - Studentga narx belgilash"])
class StudentPricingViewSet(viewsets.ModelViewSet):
    queryset         = StudentPricing.objects.all()
    serializer_class = StudentPricingSerializer

@extend_schema(tags=["StudentBalances - Studentning balansini ko'rish"])
class StudentBalancesViewSet(viewsets.ModelViewSet):
    queryset         = StudentBalances.objects.all()
    serializer_class = StudentBalancesSerializer

@extend_schema(tags=["StudentTarnsactions - Student to'lovlarni tarixi"])
class StudentTarnsactionsViewSet(viewsets.ModelViewSet):
    queryset         = StudentTarnsactions.objects.all()
    serializer_class = StudentTarnsactionsSerializer

@extend_schema(tags=["LeaveReason - Studentlarning chiqib ketish sabablari"])
class LeaveReasonViewSet(viewsets.ModelViewSet):
    queryset         = LeaveReason.objects.all()
    serializer_class = LeaveReasonSerializer

@extend_schema(tags=["StudentGroupLeaves - Studentni guruhdan chiqib ketish"])
class StudentGroupLeavesViewSet(viewsets.ModelViewSet):
    queryset         = StudentGroupLeaves.objects.all()
    serializer_class = StudentGroupLeavesSerializer

@extend_schema(tags=["StudentFreezes - Studentdarsni muzlatish"])
class StudentFreezesViewSet(viewsets.ModelViewSet):
    queryset         = StudentFreezes.objects.all()
    serializer_class = StudentFreezesSerializer

@extend_schema(tags=["StudentBalanceHistory - Student balansi tarixi"])
class StudentBalanceHistoryViewSet(viewsets.ModelViewSet):
    queryset         = StudentBalanceHistory.objects.all()
    serializer_class = StudentBalanceHistorySerializer

@extend_schema(tags=["Attendence - Yo'qlama"])
class AttendenceViewSet(viewsets.ModelViewSet):
    queryset         = Attendence.objects.all()
    serializer_class = AttendenceSerializer


# ── Group ────────────────────────────────────────────────────────

@extend_schema(tags=["Room - Xona"])
class RoomViewSet(viewsets.ModelViewSet):
    queryset         = Room.objects.all()
    serializer_class = RoomSerializer

@extend_schema(tags=["Course - Kurslar"])
class CourseViewSet(viewsets.ModelViewSet):
    queryset         = Course.objects.all()
    serializer_class = CourseSerializer

@extend_schema(tags=["Group - Guruh"])
class GroupViewSet(viewsets.ModelViewSet):
    queryset         = Group.objects.all()
    serializer_class = GroupSerializer

@extend_schema(tags=["GroupTeacher - Guruhlarni o'qituvchilarga bog'lash"])
class GroupTeacherViewSet(viewsets.ModelViewSet):
    queryset         = GroupTeacher.objects.all()
    serializer_class = GroupTeacherSerializer


# ── Lesson ───────────────────────────────────────────────────────

@extend_schema(tags=["LessonTime - Dars vaqtlari"])
class LessonTimeViewSet(viewsets.ModelViewSet):
    serializer_class = LessonTimeSerializer

    def get_queryset(self):
        qs   = LessonTime.objects.all()
        name = self.request.query_params.get('name')
        code = self.request.query_params.get('code')
        if name:
            qs = qs.filter(name__icontains=name)
        if code:
            qs = qs.filter(code__icontains=code)
        return qs

    def list(self, request, *args, **kwargs):
        qs         = self.get_queryset()
        serializer = self.get_serializer(qs, many=True)
        return Response({'count': qs.count(), 'results': serializer.data})

    def retrieve(self, request, *args, **kwargs):
        instance   = self.get_object()
        serializer = self.get_serializer(instance)
        return Response(serializer.data)

@extend_schema(tags=["LessonSchedule - Dars jadvali ro'yxati"])
class LessonScheduleViewSet(viewsets.ModelViewSet):
    queryset         = LessonSchedule.objects.all()
    serializer_class = LessonScheduleSerializer

@extend_schema(tags=["ExamsView - Imtihonlar jadvali"])
class ExamsViewSet(viewsets.ModelViewSet):
    queryset         = Exams.objects.all()
    serializer_class = ExamsSerializer

@extend_schema(tags=["ExamResults - Imtihonlar natijalari"])
class ExamResultsViewSet(viewsets.ModelViewSet):
    queryset         = ExamResults.objects.all()
    serializer_class = ExamResultsSerializer


# ── Teacher ──────────────────────────────────────────────────────

@extend_schema(tags=["TeacherSalaryRules - O'qituvchi oylik qoidalari"])
class TeacherSalaryRulesViewSet(viewsets.ModelViewSet):
    queryset         = TeacherSalaryRules.objects.all()
    serializer_class = TeacherSalaryRulesSerializer

@extend_schema(tags=["TeacherSalaryPayments - O'qituvchilarga oylik berish"])
class TeacherSalaryPaymentsViewSet(viewsets.ModelViewSet):
    queryset         = TeacherSalaryPayments.objects.all()
    serializer_class = TeacherSalaryPaymentsSerializer

@extend_schema(tags=["TeacherSalaryCalculations - O'qituvchi oyliklarini hisoblash"])
class TeacherSalaryCalculationsViewSet(viewsets.ModelViewSet):
    queryset         = TeacherSalaryCalculations.objects.all()
    serializer_class = TeacherSalaryCalculationsSerializer


# ════════════════════════════════════════════════════════════════
#  GROUP VIEWS
# ════════════════════════════════════════════════════════════════

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def group_list_create(request):
    if request.method == 'GET':
        status_filter = request.query_params.get('status')
        course_id     = request.query_params.get('course')
        room_id       = request.query_params.get('room')
        search        = request.query_params.get('search')

        groups = Group.objects.all().select_related('course', 'room')
        if status_filter:
            groups = groups.filter(status=status_filter)
        if course_id:
            groups = groups.filter(course_id=course_id)
        if room_id:
            groups = groups.filter(room_id=room_id)
        if search:
            groups = groups.filter(name__icontains=search)

        from .serializers import GroupListSerializer
        return Response({
            'success': True,
            'count':   groups.count(),
            'data':    GroupListSerializer(groups, many=True).data,
        })

    # ── POST ──
    serializer = GroupCreateUpdateSerializer(data=request.data)
    if serializer.is_valid():
        try:
            group = serializer.save()

            # ── AuditLog ──
            _log('group', group.id, 'create', None, {
                'name':       group.name,
                'course_id':  str(group.course_id),
                'status':     group.status,
                'start_date': str(group.start_date),
                'end_date':   str(group.end_date),
            }, request.user)

            from .serializers import GroupDetailSerializer
            return Response({
                'success': True,
                'message': 'Guruh muvaffaqiyatli yaratildi.',
                'data':    GroupDetailSerializer(group).data,
            }, status=status.HTTP_201_CREATED)

        except ValidationError as e:
            return Response({
                'success': False,
                'message': 'Validatsiya xatosi.',
                'errors':  e.message_dict if hasattr(e, 'message_dict') else str(e),
            }, status=status.HTTP_400_BAD_REQUEST)

    return Response({'success': False, 'message': 'Ma\'lumotlar noto\'g\'ri.',
                     'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET', 'PUT', 'PATCH', 'DELETE'])
@permission_classes([IsAuthenticated])
def group_detail_update_delete(request, pk):
    group = get_object_or_404(Group, pk=pk)
    from .serializers import GroupDetailSerializer

    if request.method == 'GET':
        return Response({'success': True, 'data': GroupDetailSerializer(group).data})

    elif request.method in ['PUT', 'PATCH']:
        # ── Eski ma'lumotni saqlab qo'yamiz ──
        old_data = {
            'name':     group.name,
            'status':   group.status,
            'room_id':  str(group.room_id),
            'end_date': str(group.end_date),
        }

        partial    = request.method == 'PATCH'
        serializer = GroupCreateUpdateSerializer(group, data=request.data, partial=partial)

        if serializer.is_valid():
            try:
                group = serializer.save()

                # ── AuditLog ──
                _log('group', group.id, 'update', old_data, {
                    'name':     group.name,
                    'status':   group.status,
                    'room_id':  str(group.room_id),
                    'end_date': str(group.end_date),
                }, request.user)

                return Response({
                    'success': True,
                    'message': 'Guruh muvaffaqiyatli yangilandi.',
                    'data':    GroupDetailSerializer(group).data,
                })

            except ValidationError as e:
                return Response({
                    'success': False,
                    'message': 'Validatsiya xatosi.',
                    'errors':  e.message_dict if hasattr(e, 'message_dict') else str(e),
                }, status=status.HTTP_400_BAD_REQUEST)

        return Response({'success': False, 'message': 'Ma\'lumotlar noto\'g\'ri.',
                         'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

    elif request.method == 'DELETE':
        old_status   = group.status
        group.status = 'archived'
        group.save()

        # ── AuditLog ──
        _log('group', group.id, 'update',
             {'status': old_status},
             {'status': 'archived', 'action': 'arxivlandi'},
             request.user)

        return Response({'success': True, 'message': 'Guruh arxivlandi.'})
# ════════════════════════════════════════════════════════════════
#  GROUP TEACHER VIEWS
# ════════════════════════════════════════════════════════════════

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def group_teacher_list_create(request, group_pk):
    group = get_object_or_404(Group, pk=group_pk)

    if request.method == 'GET':
        teachers   = group.group.all().select_related('teacher', 'group')
        serializer = GroupTeacherSerializer(teachers, many=True)
        return Response({'success': True, 'count': teachers.count(), 'data': serializer.data})

    # ── POST ──
    data          = request.data.copy()
    data['group'] = group.pk
    serializer    = GroupTeacherSerializer(data=data)

    if serializer.is_valid():
        try:
            gt = serializer.save()

            # ── AuditLog ──
            _log('group', group.id, 'update', None, {
                'action':     "O'qituvchi qo'shildi",
                'teacher_id': str(gt.teacher_id),
                'start_date': str(gt.start_date),
            }, request.user)

            return Response({'success': True, 'message': "O'qituvchi muvaffaqiyatli qo'shildi.",
                             'data': serializer.data}, status=status.HTTP_201_CREATED)

        except ValidationError as e:
            return Response({'success': False, 'message': 'Validatsiya xatosi.',
                             'errors': e.message_dict if hasattr(e, 'message_dict') else str(e)},
                            status=status.HTTP_400_BAD_REQUEST)

    return Response({'success': False, 'message': "Ma'lumotlar noto'g'ri.",
                     'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET', 'PUT', 'PATCH', 'DELETE'])
@permission_classes([IsAuthenticated])
def group_teacher_detail_update_delete(request, group_pk, pk):
    group         = get_object_or_404(Group, pk=group_pk)
    group_teacher = get_object_or_404(GroupTeacher, pk=pk, group=group)

    if request.method == 'GET':
        return Response({'success': True, 'data': GroupTeacherSerializer(group_teacher).data})

    elif request.method in ['PUT', 'PATCH']:
        partial    = request.method == 'PATCH'
        serializer = GroupTeacherSerializer(group_teacher, data=request.data, partial=partial)

        if serializer.is_valid():
            try:
                serializer.save()
                return Response({'success': True, 'message': "Ma'lumot muvaffaqiyatli yangilandi.",
                                 'data': serializer.data})
            except ValidationError as e:
                return Response({'success': False, 'message': 'Validatsiya xatosi.',
                                 'errors': e.message_dict if hasattr(e, 'message_dict') else str(e)},
                                status=status.HTTP_400_BAD_REQUEST)

        return Response({'success': False, 'message': "Ma'lumotlar noto'g'ri.",
                         'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

    elif request.method == 'DELETE':
        if group.group.count() <= 1:
            return Response({'success': False,
                             'message': "Guruhda kamida bitta o'qituvchi bo'lishi kerak."},
                            status=status.HTTP_400_BAD_REQUEST)

        teacher_id = str(group_teacher.teacher_id)
        group_teacher.delete()

        # ── AuditLog ──
        _log('group', group.id, 'update',
             {'teacher_id': teacher_id},
             {'action': "O'qituvchi guruhdan o'chirildi", 'teacher_id': teacher_id},
             request.user)

        return Response({'success': True, 'message': "O'qituvchi guruhdan o'chirildi."})


# ════════════════════════════════════════════════════════════════
#  HELPER VIEWS
# ════════════════════════════════════════════════════════════════

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def course_list(request):
    courses = Course.objects.all()
    return Response({'success': True, 'count': courses.count(),
                     'data': CourseSerializer(courses, many=True).data})


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def room_list(request):
    rooms = Room.objects.all()
    return Response({'success': True, 'count': rooms.count(),
                     'data': RoomSerializer(rooms, many=True).data})


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def available_rooms(request):
    lesson_id  = request.query_params.get('lesson')
    start_date = request.query_params.get('start_date')
    end_date   = request.query_params.get('end_date')

    if not all([lesson_id, start_date, end_date]):
        return Response({'success': False,
                         'message': 'lesson, start_date va end_date parametrlari majburiy.'},
                        status=status.HTTP_400_BAD_REQUEST)

    occupied = Group.objects.filter(
        course__lesson_id=lesson_id, status='active'
    ).filter(
        Q(start_date__lte=end_date) & Q(end_date__gte=start_date)
    ).values_list('room_id', flat=True)

    available = Room.objects.exclude(id__in=occupied)
    return Response({'success': True, 'count': available.count(),
                     'data': RoomSerializer(available, many=True).data})


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def group_statistics(request):
    return Response({'success': True, 'data': {
        'total':    Group.objects.count(),
        'active':   Group.objects.filter(status='active').count(),
        'expired':  Group.objects.filter(status='expried').count(),
        'archived': Group.objects.filter(status='archived').count(),
    }})


# ════════════════════════════════════════════════════════════════
#  GURUH MA'LUMOTLARI
# ════════════════════════════════════════════════════════════════

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def group_full_info(request, pk):
    from .serializers import GroupFullInfoSerializer
    group = get_object_or_404(Group, pk=pk)
    return Response({'success': True, 'data': GroupFullInfoSerializer(group).data})


# ════════════════════════════════════════════════════════════════
#  DAVOMAT
# ════════════════════════════════════════════════════════════════

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def attendance_list_create(request, group_pk):
    from .serializers import AttendanceSerializer
    group = get_object_or_404(Group, pk=group_pk)

    if request.method == 'GET':
        date_q      = request.query_params.get('date')
        attendances = Attendence.objects.filter(student_group__group=group)
        if date_q:
            attendances = attendances.filter(lesson_date=date_q)
        return Response({'success': True, 'data': AttendanceSerializer(attendances, many=True).data})

    serializer = AttendanceSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save(marked_by=request.user.employee)
        return Response({'success': True, 'data': serializer.data}, status=status.HTTP_201_CREATED)
    return Response({'success': False, 'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)


# ════════════════════════════════════════════════════════════════
#  O'QUVCHILAR RO'YXATI
# ════════════════════════════════════════════════════════════════

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def students_list(request, group_pk):
    from .serializers import StudentListSerializer
    group         = get_object_or_404(Group, pk=group_pk)
    status_filter = request.query_params.get('status')
    search        = request.query_params.get('search')

    students = StudentGroup.objects.filter(group=group)
    if status_filter == 'active':
        students = students.filter(left_at__isnull=True)
    elif status_filter == 'left':
        students = students.filter(left_at__isnull=False)
    if search:
        students = students.filter(student__full_name__icontains=search)

    return Response({'success': True, 'data': StudentListSerializer(students, many=True).data})


# ════════════════════════════════════════════════════════════════
#  GURUHGA TALABA QO'SHISH / O'CHIRISH
# ════════════════════════════════════════════════════════════════

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def add_student(request, group_pk):
    from .serializers import AddStudentSerializer
    group         = get_object_or_404(Group, pk=group_pk)
    data          = request.data.copy()
    data['group'] = group.pk

    serializer = AddStudentSerializer(data=data)
    if serializer.is_valid():
        sg = serializer.save()

        # ── AuditLog ──
        _log('student', sg.student_id, 'update', None, {
            'action':    "Guruhga qo'shildi",
            'group_id':  str(group.id),
            'group':     group.name,
            'joined_at': str(sg.joined_at),
        }, request.user)

        return Response({'success': True, 'data': serializer.data}, status=status.HTTP_201_CREATED)
    return Response({'success': False, 'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)


@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def remove_student(request, group_pk, student_group_pk):
    sg = get_object_or_404(StudentGroup, pk=student_group_pk, group_id=group_pk)

    # ── AuditLog ──
    _log('student', sg.student_id, 'update',
         {'left_at': None},
         {'action':   "Guruhdan chiqarildi",
          'group_id': str(sg.group_id),
          'group':    sg.group.name,
          'left_at':  str(timezone.now().date())},
         request.user)

    sg.left_at = timezone.now().date()
    sg.save()
    return Response({'success': True, 'message': "Talaba o'chirildi"})


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def send_sms(request, group_pk):
    from .serializers import SMSSerializer
    group      = get_object_or_404(Group, pk=group_pk)
    serializer = SMSSerializer(data=request.data)

    if serializer.is_valid():
        message  = serializer.validated_data['message']
        send_to  = serializer.validated_data['send_to']
        phones   = []
        students = StudentGroup.objects.filter(group=group, left_at__isnull=True)

        for sg in students:
            if send_to in ['all', 'students']:
                phones.append(sg.student.phone_number)
            if send_to in ['all', 'parents'] and sg.student.parent_phone:
                phones.append(sg.student.parent_phone)

        # TODO: SMS provider integration
        return Response({'success': True, 'message': f'{len(phones)} ta SMS yuborildi'})

    return Response({'success': False, 'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)


# ════════════════════════════════════════════════════════════════
#  CHEGIRMALAR
# ════════════════════════════════════════════════════════════════

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def discount_list_create(request):
    from .serializers import DiscountSerializer
    if request.method == 'GET':
        student_id = request.query_params.get('student')
        discounts  = StudentPricing.objects.all()
        if student_id:
            discounts = discounts.filter(student_id=student_id)
        return Response({'success': True, 'data': DiscountSerializer(discounts, many=True).data})

    serializer = DiscountSerializer(data=request.data)
    if serializer.is_valid():
        discount = serializer.save(created_by=request.user.employee)

        # ── AuditLog ──
        _log('student', discount.student_id, 'update', None, {
            'action':         'Chegirma berildi',
            'price_override': str(discount.price_override),
            'course_id':      str(discount.course_id),
            'start_date':     str(discount.start_date),
            'end_date':       str(discount.end_date),
        }, request.user)

        return Response({'success': True, 'data': serializer.data}, status=status.HTTP_201_CREATED)
    return Response({'success': False, 'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET', 'PUT', 'DELETE'])
@permission_classes([IsAuthenticated])
def discount_detail(request, pk):
    from .serializers import DiscountSerializer
    discount = get_object_or_404(StudentPricing, pk=pk)

    if request.method == 'GET':
        return Response({'success': True, 'data': DiscountSerializer(discount).data})

    elif request.method == 'PUT':
        old_price  = str(discount.price_override)
        serializer = DiscountSerializer(discount, data=request.data)
        if serializer.is_valid():
            serializer.save()

            # ── AuditLog ──
            _log('student', discount.student_id, 'update',
                 {'price_override': old_price},
                 {'price_override': str(discount.price_override), 'action': 'Chegirma yangilandi'},
                 request.user)

            return Response({'success': True, 'data': serializer.data})
        return Response({'success': False, 'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

    else:  # DELETE
        related = StudentGroup.objects.filter(
            student=discount.student,
            group__course=discount.course,
            left_at__isnull=True
        )
        if related.exists():
            return Response({'success': False,
                             'message': f"Bu chegirmani o'chirish mumkin emas. "
                                        f"{related.count()} ta faol guruhda ishlatilmoqda."},
                            status=status.HTTP_400_BAD_REQUEST)

        # ── AuditLog ──
        _log('student', discount.student_id, 'update',
             {'price_override': str(discount.price_override), 'course_id': str(discount.course_id)},
             {'action': "Chegirma o'chirildi"},
             request.user)

        discount.delete()
        return Response({'success': True, 'message': "Chegirma o'chirildi"})


# ════════════════════════════════════════════════════════════════
#  IMTIHONLAR
# ════════════════════════════════════════════════════════════════

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def exam_list_create(request):
    from .serializers import ExamSerializer
    if request.method == 'GET':
        group_id = request.query_params.get('group')
        exams    = Exams.objects.all()
        if group_id:
            exams = exams.filter(group_id=group_id)
        return Response({'success': True, 'data': ExamSerializer(exams, many=True).data})

    serializer = ExamSerializer(data=request.data)
    if serializer.is_valid():
        exam = serializer.save(created_by=request.user.employee)

        # ── AuditLog ──
        _log('other', exam.id, 'create', None, {
            'action':    'Imtihon yaratildi',
            'title':     exam.title,
            'group_id':  str(exam.group_id),
            'exam_date': str(exam.exam_date),
        }, request.user)

        return Response({'success': True, 'data': serializer.data}, status=status.HTTP_201_CREATED)
    return Response({'success': False, 'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET', 'PUT', 'DELETE'])
@permission_classes([IsAuthenticated])
def exam_detail(request, pk):
    from .serializers import ExamSerializer
    exam = get_object_or_404(Exams, pk=pk)

    if request.method == 'GET':
        return Response({'success': True, 'data': ExamSerializer(exam).data})

    elif request.method == 'PUT':
        old_data   = {'title': exam.title, 'exam_date': str(exam.exam_date)}
        serializer = ExamSerializer(exam, data=request.data)
        if serializer.is_valid():
            serializer.save()

            # ── AuditLog ──
            _log('other', exam.id, 'update', old_data, {
                'title':     exam.title,
                'exam_date': str(exam.exam_date),
            }, request.user)

            return Response({'success': True, 'data': serializer.data})
        return Response({'success': False, 'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

    else:  # DELETE
        results_count = ExamResults.objects.filter(exam=exam).count()
        if results_count > 0:
            return Response({'success': False,
                             'message': f"Bu imtihonni o'chirish mumkin emas. "
                                        f"{results_count} ta natija mavjud. Avval natijalarni o'chiring."},
                            status=status.HTTP_400_BAD_REQUEST)

        # ── AuditLog ──
        _log('other', exam.id, 'delete',
             {'title': exam.title, 'group_id': str(exam.group_id)},
             None, request.user)

        exam.delete()
        return Response({'success': True, 'message': "Imtihon o'chirildi"})


@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def exam_results(request, exam_pk):
    from .serializers import ExamResultSerializer
    exam = get_object_or_404(Exams, pk=exam_pk)

    if request.method == 'GET':
        results = ExamResults.objects.filter(exam=exam)
        return Response({'success': True, 'data': ExamResultSerializer(results, many=True).data})

    data         = request.data.copy()
    data['exam'] = exam.pk
    serializer   = ExamResultSerializer(data=data)
    if serializer.is_valid():
        result = serializer.save(created_by=request.user.employee)

        # ── AuditLog ──
        _log('other', exam.id, 'update', None, {
            'action':     'Imtihon natijasi kiritildi',
            'student_id': str(result.student_id),
            'score':      str(result.score),
        }, request.user)

        return Response({'success': True, 'data': serializer.data}, status=status.HTTP_201_CREATED)
    return Response({'success': False, 'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)


# ════════════════════════════════════════════════════════════════
#  GURUH TARIXI (AUDIT LOG)
# ════════════════════════════════════════════════════════════════

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def group_history(request, group_pk):
    group = get_object_or_404(Group, pk=group_pk)

    logs = AuditLog.objects.filter(
        entity_type='group',
        entity_id=group.id
    ).order_by('-created_at')

    action = request.query_params.get('action')
    if action:
        logs = logs.filter(action=action)

    date_val = request.query_params.get('date')
    if date_val:
        parsed = parse_date(date_val)
        if parsed:
            logs = logs.filter(created_at__date=parsed)

    date_from = request.query_params.get('from')
    date_to   = request.query_params.get('to')
    if date_from:
        logs = logs.filter(created_at__date__gte=date_from)
    if date_to:
        logs = logs.filter(created_at__date__lte=date_to)

    data = [{
        'id':           str(log.id),
        'action':       log.action,
        'old_data':     log.old_data,
        'new_data':     log.new_action,
        'performed_by': (
            f"{log.performed_by.user.full_name}"
            if log.performed_by and log.performed_by.user else None
        ),
        'created_at': log.created_at,
    } for log in logs]

    return Response(data)

# ════════════════════════════════════════════════════════════════
#  1. O'QUVCHINI QIDIRISH
# ════════════════════════════════════════════════════════════════

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def student_search(request):
    query = request.query_params.get('q', '').strip()

    if not query:
        return Response({
            'success': False,
            'message': "Qidiruv so'zi kiritilmagan. q parametrini kiriting."
        }, status=status.HTTP_400_BAD_REQUEST)

    students = Student.objects.filter(
        Q(full_name__icontains=query) |
        Q(phone_number__icontains=query) |
        Q(phone_number2__icontains=query)
    ).distinct()[:20]

    return Response({
        'success': True,
        'count':   students.count(),
        'query':   query,
        'data':    StudentSearchSerializer(students, many=True).data,
    })


# ════════════════════════════════════════════════════════════════
#  2. GURUHGA TALABA QO'SHISH
# ════════════════════════════════════════════════════════════════

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def add_student_to_group(request, group_pk):
    group         = get_object_or_404(Group, pk=group_pk)
    data          = request.data.copy()
    data['group'] = group.pk

    serializer = AddStudentToGroupSerializer(data=data)

    if serializer.is_valid():
        student_group = serializer.save()

        # ── AuditLog ──
        _log('student', student_group.student_id, 'update', None, {
            'action':    "Guruhga qo'shildi",
            'group_id':  str(group.id),
            'group':     group.name,
            'joined_at': str(student_group.joined_at),
        }, request.user)

        return Response({
            'success': True,
            'message': "Talaba muvaffaqiyatli guruhga qo'shildi.",
            'data':    serializer.data,
        }, status=status.HTTP_201_CREATED)

    return Response({
        'success': False,
        'message': "Ma'lumotlar noto'g'ri.",
        'errors':  serializer.errors,
    }, status=status.HTTP_400_BAD_REQUEST)


# ════════════════════════════════════════════════════════════════
#  3. TALABANI FAOLLASHTIRISH
# ════════════════════════════════════════════════════════════════

@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def activate_student(request, student_pk):
    student    = get_object_or_404(Student, pk=student_pk)
    serializer = ActivateStudentSerializer(data=request.data)

    if serializer.is_valid():
        try:
            result = serializer.activate_student(
                student=student,
                validated_data=serializer.validated_data,
                user=request.user.employee,
            )

            # ── AuditLog ──
            _log('payment', student.id, 'create', None, {
                'action':      'Talaba faollashtirildi',
                'old_balance': result.get('old_balance'),
                'added_amount': result.get('added_amount'),
                'new_balance': result.get('new_balance'),
            }, request.user)

            return Response({
                'success': True,
                'message': 'Talaba muvaffaqiyatli faollashtirildi.',
                'data':    result,
            })

        except Exception as e:
            return Response({
                'success': False,
                'message': f'Xatolik yuz berdi: {str(e)}',
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    return Response({
        'success': False,
        'message': "Ma'lumotlar noto'g'ri.",
        'errors':  serializer.errors,
    }, status=status.HTTP_400_BAD_REQUEST)


# ════════════════════════════════════════════════════════════════
#  4. TALABA BALANSI VA HOLATI
# ════════════════════════════════════════════════════════════════

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def student_balance_status(request, student_pk):
    student = get_object_or_404(Student, pk=student_pk)

    try:
        student_balance = student.Sbalances_student.first()
        balance = float(student_balance.balance) if student_balance else 0.0
    except Exception:
        balance = 0.0

    active_groups      = student.Sgroup_student.filter(left_at__isnull=True).count()
    transactions_count = student.Strans_student.count()
    student_status     = 'active' if balance > 0 else 'inactive'

    return Response({
        'success': True,
        'data': {
            'student_id':         student.id,
            'full_name':          student.full_name,
            'phone_number':       student.phone_number,
            'balance':            balance,
            'status':             student_status,
            'active_groups':      active_groups,
            'transactions_count': transactions_count,
        },
    })


# ════════════════════════════════════════════════════════════════
#  ONLAYN DARSLAR
# ════════════════════════════════════════════════════════════════

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def group_online_lessons(request, group_pk):
    group = get_object_or_404(Group, pk=group_pk)

    if request.method == 'GET':
        is_published = request.query_params.get('is_published')
        content_type = request.query_params.get('content_type')
        lessons      = OnlineLesson.objects.filter(group=group)

        if is_published is not None:
            lessons = lessons.filter(is_published=is_published.lower() in ['true', '1'])
        if content_type:
            lessons = lessons.filter(content_type=content_type)

        return Response({
            'success': True,
            'count':   lessons.count(),
            'data':    OnlineLessonListSerializer(lessons, many=True, context={'request': request}).data,
        })

    # ── POST ──
    data          = request.data.copy()
    data['group'] = group.pk
    serializer    = OnlineLessonCreateUpdateSerializer(data=data, context={'request': request})

    if serializer.is_valid():
        lesson = serializer.save(created_by=request.user.employee)

        # ── AuditLog ──
        _log('other', lesson.id, 'create', None, {
            'action':    'Onlayn dars yaratildi',
            'title':     lesson.title,
            'group_id':  str(group.id),
            'group':     group.name,
        }, request.user)

        return Response({
            'success': True,
            'message': 'Dars yaratildi.',
            'data':    serializer.data,
        }, status=status.HTTP_201_CREATED)

    return Response({'success': False, 'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET', 'PATCH', 'DELETE'])
@permission_classes([IsAuthenticated])
def online_lesson_detail(request, pk):
    lesson = get_object_or_404(OnlineLesson, pk=pk)

    if request.method == 'GET':
        return Response({
            'success': True,
            'data':    OnlineLessonDetailSerializer(lesson, context={'request': request}).data,
        })

    elif request.method == 'PATCH':
        old_title  = lesson.title
        serializer = OnlineLessonCreateUpdateSerializer(
            lesson, data=request.data, partial=True, context={'request': request}
        )
        if serializer.is_valid():
            serializer.save()

            # ── AuditLog ──
            _log('other', lesson.id, 'update',
                 {'title': old_title},
                 {'title': lesson.title, 'action': 'Onlayn dars yangilandi'},
                 request.user)

            return Response({'success': True, 'message': 'Dars yangilandi.', 'data': serializer.data})

        return Response({'success': False, 'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

    else:  # DELETE
        # ── AuditLog ──
        _log('other', lesson.id, 'delete',
             {'title': lesson.title, 'group_id': str(lesson.group_id)},
             None, request.user)

        lesson.delete()
        return Response({'success': True, 'message': "Dars o'chirildi."})


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def publish_lesson(request, pk):
    lesson             = get_object_or_404(OnlineLesson, pk=pk)
    lesson.is_published = True
    lesson.save()
    return Response({
        'success': True,
        'message': 'Dars nashr qilindi.',
        'data':    OnlineLessonDetailSerializer(lesson, context={'request': request}).data,
    })


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def unpublish_lesson(request, pk):
    lesson             = get_object_or_404(OnlineLesson, pk=pk)
    lesson.is_published = False
    lesson.save()
    return Response({
        'success': True,
        'message': 'Dars nashrdan olindi.',
        'data':    OnlineLessonDetailSerializer(lesson, context={'request': request}).data,
    })


# ════════════════════════════════════════════════════════════════
#  STUDENT DISCOUNT VIEWS
# ════════════════════════════════════════════════════════════════

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def create_student_discount(request, group_pk, student_group_pk):
    group         = get_object_or_404(Group, pk=group_pk)
    student_group = get_object_or_404(StudentGroup, pk=student_group_pk, group=group)
    student       = student_group.student

    serializer = CreateStudentDiscountSerializer(data=request.data)

    if serializer.is_valid():
        try:
            discount = serializer.create_discount(
                student=student,
                group=group,
                validated_data=serializer.validated_data,
                user=request.user.employee,
            )

            # ── AuditLog ──
            _log('student', student.id, 'update', None, {
                'action':         'Chegirma yaratildi',
                'price_override': str(discount.price_override),
                'course_id':      str(discount.course_id),
                'start_date':     str(discount.start_date),
                'end_date':       str(discount.end_date),
            }, request.user)

            return Response({
                'success': True,
                'message': 'Chegirma muvaffaqiyatli yaratildi.',
                'data':    StudentDiscountSerializer(discount, context={'request': request}).data,
            }, status=status.HTTP_201_CREATED)

        except Exception as e:
            return Response({'success': False, 'message': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    return Response({'success': False, 'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)


@api_view(['PATCH'])
@permission_classes([IsAuthenticated])
def update_student_discount(request, group_pk, student_group_pk):
    group         = get_object_or_404(Group, pk=group_pk)
    student_group = get_object_or_404(StudentGroup, pk=student_group_pk, group=group)
    student       = student_group.student
    course        = group.course
    today         = timezone.now().date()

    try:
        discount = StudentPricing.objects.get(
            student=student, course=course,
            start_date__lte=today, end_date__gte=today,
        )
    except StudentPricing.DoesNotExist:
        return Response({
            'success': False,
            'message': "Bu o'quvchi uchun aktiv chegirma topilmadi.",
        }, status=status.HTTP_404_NOT_FOUND)

    old_price  = str(discount.price_override)
    serializer = StudentDiscountSerializer(discount, data=request.data, partial=True)

    if serializer.is_valid():
        serializer.save()

        # ── AuditLog ──
        _log('student', student.id, 'update',
             {'price_override': old_price},
             {'price_override': str(discount.price_override), 'action': 'Chegirma yangilandi'},
             request.user)

        return Response({'success': True, 'message': 'Chegirma yangilandi.', 'data': serializer.data})

    return Response({'success': False, 'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)


@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def delete_student_discount(request, group_pk, student_group_pk):
    group         = get_object_or_404(Group, pk=group_pk)
    student_group = get_object_or_404(StudentGroup, pk=student_group_pk, group=group)
    student       = student_group.student
    course        = group.course
    today         = timezone.now().date()

    try:
        discount = StudentPricing.objects.get(
            student=student, course=course,
            start_date__lte=today, end_date__gte=today,
        )
    except StudentPricing.DoesNotExist:
        return Response({
            'success': False,
            'message': "Bu o'quvchi uchun aktiv chegirma topilmadi.",
        }, status=status.HTTP_404_NOT_FOUND)

    old_price        = str(discount.price_override)
    discount.end_date = today
    discount.save()

    # ── AuditLog ──
    _log('student', student.id, 'update',
         {'price_override': old_price, 'end_date': str(today)},
         {'action': 'Chegirma bekor qilindi', 'end_date': str(today)},
         request.user)

    return Response({'success': True, 'message': 'Chegirma bekor qilindi.'})


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_student_discount(request, group_pk, student_group_pk):
    group         = get_object_or_404(Group, pk=group_pk)
    student_group = get_object_or_404(StudentGroup, pk=student_group_pk, group=group)
    student       = student_group.student
    course        = group.course
    today         = timezone.now().date()

    discounts = StudentPricing.objects.filter(student=student, course=course).order_by('-created_at')

    if not discounts.exists():
        return Response({'success': True, 'message': 'Chegirma topilmadi.', 'data': None})

    active_discount = discounts.filter(start_date__lte=today, end_date__gte=today).first()

    return Response({
        'success':   True,
        'data':      StudentDiscountSerializer(active_discount or discounts.first(),
                                               context={'request': request}).data,
        'is_active': active_discount is not None,
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def calculate_payment_difference(request, group_pk, student_group_pk):
    group         = get_object_or_404(Group, pk=group_pk)
    student_group = get_object_or_404(StudentGroup, pk=student_group_pk, group=group)
    student       = student_group.student
    course        = group.course
    today         = timezone.now().date()

    try:
        discount = StudentPricing.objects.get(
            student=student, course=course,
            start_date__lte=today, end_date__gte=today,
        )
    except StudentPricing.DoesNotExist:
        return Response({'success': False, 'message': 'Aktiv chegirma topilmadi.'},
                        status=status.HTTP_404_NOT_FOUND)

    original_price = float(course.monthly_price)
    discount_price = float(discount.price_override)
    difference     = original_price - discount_price

    from dateutil.relativedelta import relativedelta
    months_remaining = 0
    if group.end_date > today:
        delta            = relativedelta(group.end_date, today)
        months_remaining = delta.years * 12 + delta.months

    return Response({
        'success': True,
        'data': {
            'original_price':       original_price,
            'discount_price':       discount_price,
            'difference':           difference,
            'discount_percentage':  round((difference / original_price) * 100, 2),
            'months_remaining':     months_remaining,
            'total_savings':        difference * months_remaining,
            'group_end_date':       group.end_date,
        },
    })


# ════════════════════════════════════════════════════════════════
#  DARS SANASI BOSHQARUVI
# ════════════════════════════════════════════════════════════════

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def group_lessons(request, group_pk):
    group    = get_object_or_404(Group, pk=group_pk)
    day_type = request.query_params.get('day_type')
    lessons  = LessonSchedule.objects.filter(group=group)

    if day_type:
        lessons = lessons.filter(day_type=day_type)
    lessons = lessons.order_by('day_type', 'start_time')

    return Response({
        'success': True,
        'count':   lessons.count(),
        'data':    LessonScheduleListSerializer(lessons, many=True).data,
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def lesson_detail(request, group_pk, lesson_pk):
    group  = get_object_or_404(Group, pk=group_pk)
    lesson = get_object_or_404(LessonSchedule, pk=lesson_pk, group=group)
    return Response({'success': True, 'data': LessonScheduleDetailSerializer(lesson).data})


@api_view(['PATCH'])
@permission_classes([IsAuthenticated])
def move_lesson_date(request, group_pk, lesson_pk):
    group      = get_object_or_404(Group, pk=group_pk)
    lesson     = get_object_or_404(LessonSchedule, pk=lesson_pk, group=group)
    serializer = MoveLessonDateSerializer(data=request.data)

    if serializer.is_valid():
        try:
            old_start    = lesson.start_time
            old_end      = lesson.end_time
            old_day_type = lesson.day_type

            updated_lesson = serializer.move_lesson(lesson, serializer.validated_data)

            # ── AuditLog ──
            _log('other', lesson.id, 'update', {
                'day_type':   old_day_type,
                'start_time': str(old_start),
                'end_time':   str(old_end),
            }, {
                'action':     'Dars vaqti ko\'chirildi',
                'day_type':   updated_lesson.day_type,
                'start_time': str(updated_lesson.start_time),
                'end_time':   str(updated_lesson.end_time),
                'group_id':   str(group.id),
            }, request.user)

            return Response({
                'success':      True,
                'message':      "Dars sanasi muvaffaqiyatli ko'chirildi.",
                'old_schedule': {
                    'day_type':   old_day_type,
                    'start_time': str(old_start),
                    'end_time':   str(old_end),
                },
                'new_schedule': {
                    'day_type':   updated_lesson.day_type,
                    'start_time': str(updated_lesson.start_time),
                    'end_time':   str(updated_lesson.end_time),
                },
                'data': LessonScheduleDetailSerializer(updated_lesson).data,
            })

        except Exception as e:
            return Response({'success': False, 'message': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    return Response({'success': False, 'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def check_time_availability(request, group_pk):
    group             = get_object_or_404(Group, pk=group_pk)
    start_time        = request.data.get('start_time')
    end_time          = request.data.get('end_time')
    day_type          = request.data.get('day_type')
    exclude_lesson_id = request.data.get('exclude_lesson_id')

    if not all([start_time, end_time, day_type]):
        return Response({'success': False, 'message': 'start_time, end_time va day_type majburiy.'},
                        status=status.HTTP_400_BAD_REQUEST)

    # Xona to'qnashuvi
    room_conflicts = LessonSchedule.objects.filter(
        group__room=group.room, day_type=day_type
    ).filter(Q(start_time__lt=end_time) & Q(end_time__gt=start_time))
    if exclude_lesson_id:
        room_conflicts = room_conflicts.exclude(pk=exclude_lesson_id)

    # O'qituvchi to'qnashuvi
    teachers          = GroupTeacher.objects.filter(group=group)
    teacher_conflicts = []
    for gt in teachers:
        conflicts = LessonSchedule.objects.filter(
            group__group__teacher=gt.teacher, day_type=day_type
        ).filter(Q(start_time__lt=end_time) & Q(end_time__gt=start_time))
        if exclude_lesson_id:
            conflicts = conflicts.exclude(pk=exclude_lesson_id)
        if conflicts.exists():
            teacher_conflicts.append({
                'teacher':        f"{gt.teacher.user.full_name}",
                'conflict_group': conflicts.first().group.name,
            })

    return Response({
        'success':           True,
        'is_available':      not room_conflicts.exists() and not teacher_conflicts,
        'room_conflicts':    [{'group': c.group.name, 'start_time': str(c.start_time),
                               'end_time': str(c.end_time)} for c in room_conflicts],
        'teacher_conflicts': teacher_conflicts,
    })


# ════════════════════════════════════════════════════════════════
#  GURUHNI TARK ETGAN TALABALAR HISOBOTI
# ════════════════════════════════════════════════════════════════

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def student_leave_report(request):
    filter_serializer = LeaveReportFilterSerializer(data=request.data)

    if not filter_serializer.is_valid():
        return Response({'success': False, 'errors': filter_serializer.errors},
                        status=status.HTTP_400_BAD_REQUEST)

    filters = filter_serializer.validated_data

    leaves = StudentGroupLeaves.objects.select_related(
        'student', 'group', 'group__course', 'leave_reason', 'created_by'
    ).all()

    if filters.get('leave_date_start'):
        leaves = leaves.filter(leave_date__gte=filters['leave_date_start'])
    if filters.get('leave_date_end'):
        leaves = leaves.filter(leave_date__lte=filters['leave_date_end'])
    if filters.get('search'):
        s = filters['search']
        leaves = leaves.filter(
            Q(student__full_name__icontains=s) |
            Q(student__phone_number__icontains=s) |
            Q(student__phone_number2__icontains=s)
        )
    if filters.get('course_id'):
        leaves = leaves.filter(group__course_id=filters['course_id'])
    if filters.get('group_id'):
        leaves = leaves.filter(group_id=filters['group_id'])
    if filters.get('teacher_id'):
        leaves = leaves.filter(group__group__teacher_id=filters['teacher_id'])
    if filters.get('archived_by_id'):
        leaves = leaves.filter(created_by_id=filters['archived_by_id'])
    if filters.get('leave_reason_id'):
        leaves = leaves.filter(leave_reason_id=filters['leave_reason_id'])

    results = []
    for leave in leaves:
        teachers     = leave.group.group.all()
        teacher_names = ', '.join([
            f"{t.teacher.user.full_name}" for t in teachers
        ])
        results.append({
            'student_id':   leave.student.id,
            'full_name':    leave.student.full_name,
            'phone_number': leave.student.phone_number,
            'course_name':  leave.group.course.name if leave.group.course else '-',
            'group_name':   leave.group.name,
            'teacher_name': teacher_names or '-',
            'status':       'archived',
            'archived_by':  leave.created_by.user.full_name if leave.created_by else '-',
            'leave_date':   leave.leave_date,
            'leave_reason': leave.leave_reason.name if leave.leave_reason else '-',
            'comment':      leave.comment or '-',
        })

    return Response({
        'success': True,
        'filters': filters,
        'count':   len(results),
        'data':    StudentLeaveReportSerializer(results, many=True).data,
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def leave_reasons_list(request):
    reasons = LeaveReason.objects.filter(is_active=True).values('id', 'name')
    return Response({'success': True, 'data': list(reasons)})


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def teachers_list(request):
    teachers = Employee.objects.filter(is_active=True).select_related('user')
    data     = [{'id': t.id, 'name': t.user.full_name} for t in teachers]
    return Response({'success': True, 'data': data})


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def courses_list(request):
    courses = Course.objects.all().values('id', 'name')
    return Response({'success': True, 'data': list(courses)})

# bu darsni bekor qilish

# ==================== VIEWS ====================

# ════════════════════════════════════════════════════════════════
#  DARSNI BEKOR QILISH
# ════════════════════════════════════════════════════════════════

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def cancel_lesson(request, group_pk, lesson_pk):
    group      = get_object_or_404(Group, pk=group_pk)
    lesson     = get_object_or_404(LessonSchedule, pk=lesson_pk, group=group)
    serializer = CancelLessonSerializer(
        data=request.data,
        context={'lesson_schedule_id': lesson.id}
    )

    if serializer.is_valid():
        try:
            audit = serializer.cancel_lesson(
                group=group,
                lesson_schedule=lesson,
                validated_data=serializer.validated_data,
                user=request.user.employee,
            )

            # ── AuditLog ──
            _log('other', lesson.id, 'update', None, {
                'action':             'Dars bekor qilindi',
                'group_id':           str(group.id),
                'group':              group.name,
                'lesson_date':        audit.old_data.get('lesson_date'),
                'reason':             serializer.validated_data.get('cancellation_reason', ''),
                'refund_per_student': audit.new_action.get('refund_per_student', 0),
                'total_refunded':     audit.new_action.get('total_refunded', 0),
                'students_count':     audit.new_action.get('students_count', 0),
            }, request.user)

            return Response({
                'success': True,
                'message': 'Dars bekor qilindi.',
                'data': {
                    'id':                 audit.id,
                    'lesson_date':        audit.old_data['lesson_date'],
                    'refund_per_student': audit.new_action.get('refund_per_student', 0),
                    'total_refunded':     audit.new_action.get('total_refunded', 0),
                    'students_count':     audit.new_action.get('students_count', 0),
                },
            }, status=status.HTTP_201_CREATED)

        except Exception as e:
            return Response({'success': False, 'message': f'Xatolik: {str(e)}'},
                            status=status.HTTP_400_BAD_REQUEST)

    return Response({'success': False, 'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)


# ════════════════════════════════════════════════════════════════
#  DARSNI QAYTA TIKLASH
# ════════════════════════════════════════════════════════════════

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def restore_lesson(request, audit_pk):
    audit = get_object_or_404(
        AuditLog, pk=audit_pk,
        entity_type='cancelled_lesson', action='cancelled'
    )
    serializer = RestoreLessonSerializer(data=request.data)

    if serializer.is_valid():
        try:
            restored = serializer.restore_lesson(
                cancelled_audit=audit,
                user=request.user.employee,
            )

            # ── AuditLog ──
            _log('other', audit.id, 'update',
                 {'action': 'cancelled'},
                 {
                     'action':            'Dars qayta tiklandi',
                     'restored_at':       restored.new_action.get('restored_at'),
                     'restored_by_name':  restored.new_action.get('restored_by_name'),
                     'restore_note':      serializer.validated_data.get('restore_note', ''),
                 }, request.user)

            return Response({
                'success': True,
                'message': 'Dars qayta tiklandi. DIQQAT: Balans avtomatik qaytarilmadi!',
                'data': {
                    'id':               restored.id,
                    'restored_at':      restored.new_action.get('restored_at'),
                    'restored_by_name': restored.new_action.get('restored_by_name'),
                },
            })

        except Exception as e:
            return Response({'success': False, 'message': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    return Response({'success': False, 'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)


# ════════════════════════════════════════════════════════════════
#  BEKOR QILINGAN DARSLAR RO'YXATI
# ════════════════════════════════════════════════════════════════

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def cancelled_lessons_list(request, group_pk):
    group = get_object_or_404(Group, pk=group_pk)

    cancelled = AuditLog.objects.filter(
        entity_type='cancelled_lesson',
        action='cancelled',
    ).filter(old_data__group_id=group.id).order_by('-created_at')

    is_restored = request.query_params.get('is_restored')
    if is_restored == 'true':
        cancelled = cancelled.filter(new_action__has_key='restored_at')
    elif is_restored == 'false':
        cancelled = cancelled.exclude(new_action__has_key='restored_at')

    results = []
    for audit in cancelled:
        results.append({
            'id':               audit.id,
            'lesson_date':      audit.old_data.get('lesson_date'),
            'group_name':       audit.old_data.get('group_name'),
            'reason':           audit.new_action.get('reason', ''),
            'cancelled_by_name': audit.new_action.get('cancelled_by_name'),
            'cancelled_at':     audit.new_action.get('cancelled_at'),
            'refund_per_student': audit.new_action.get('refund_per_student', 0),
            'total_refunded':   audit.new_action.get('total_refunded', 0),
            'is_restored':      'restored_at' in audit.new_action,
            'restored_at':      audit.new_action.get('restored_at'),
            'restored_by_name': audit.new_action.get('restored_by_name'),
        })

    return Response({'success': True, 'count': len(results), 'data': results})


# ════════════════════════════════════════════════════════════════
#  CHEGIRMALI NARX BELGILASH
# ════════════════════════════════════════════════════════════════

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def set_discount_price(request, group_pk, student_group_pk):
    group         = get_object_or_404(Group, pk=group_pk)
    student_group = get_object_or_404(StudentGroup, pk=student_group_pk, group=group)
    student       = student_group.student
    course        = group.course

    serializer = DiscountPriceSerializer(data=request.data)

    if serializer.is_valid():
        with transaction.atomic():
            today = timezone.now().date()

            # 1. Eski chegirmalarni bekor qilish
            old_discounts = StudentPricing.objects.filter(
                student=student, course=course, end_date__gte=today
            )
            old_prices = [str(o.price_override) for o in old_discounts]
            for old in old_discounts:
                old.end_date = today
                old.save()

            # 2. Yangi chegirma yaratish
            new_discount = StudentPricing.objects.create(
                student=student,
                course=course,
                price_override=serializer.validated_data['price_override'],
                reason=serializer.validated_data.get('reason', ''),
                start_date=serializer.validated_data['start_date'],
                end_date=serializer.validated_data['end_date'],
                created_by=request.user.employee,
            )

            # 3. Faollashtirilganmi tekshirish
            is_active = (
                student_group.joined_at <= today and
                (student_group.left_at is None or student_group.left_at >= today)
            )
            refunded = 0

            if is_active:
                original_price = course.monthly_price
                new_price      = new_discount.price_override

                if new_price < original_price:
                    refund = original_price - new_price

                    balance, _ = StudentBalances.objects.get_or_create(
                        student=student, defaults={'balance': 0}
                    )
                    balance.balance += refund
                    balance.save()

                    StudentTarnsactions.objects.create(
                        student=student,
                        student_group=student_group,
                        group=group,
                        transaction_type='refund',
                        amount=refund,
                        payment_type='correction',
                        transaction_date=timezone.now(),
                        accepted_by=request.user.employee,
                        comment=f"Chegirma belgilandi: {original_price} → {new_price}",
                    )
                    refunded = float(refund)

            # ── AuditLog ──
            _log('student', student.id, 'update',
                 {'old_prices': old_prices},
                 {
                     'action':         'Chegirmali narx belgilandi',
                     'price_override': str(new_discount.price_override),
                     'course_id':      str(course.id),
                     'start_date':     str(new_discount.start_date),
                     'end_date':       str(new_discount.end_date),
                     'refunded':       refunded,
                     'is_active':      is_active,
                 }, request.user)

            return Response({
                'success': True,
                'message': 'Chegirmali narx belgilandi.',
                'data':    StudentDiscountSerializer(new_discount).data,
                'refund_info': {
                    'is_active':      is_active,
                    'refunded_amount': refunded,
                    'message':        f"{refunded} so'm qaytarildi" if refunded > 0 else 'Hali faollashtirilmagan',
                },
            }, status=status.HTTP_201_CREATED)

    return Response({'success': False, 'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_student_discount(request, group_pk, student_group_pk):
    group         = get_object_or_404(Group, pk=group_pk)
    student_group = get_object_or_404(StudentGroup, pk=student_group_pk, group=group)
    student       = student_group.student
    course        = group.course
    today         = timezone.now().date()

    discount = StudentPricing.objects.filter(
        student=student, course=course,
        start_date__lte=today, end_date__gte=today,
    ).first()

    if not discount:
        return Response({
            'success':        True,
            'message':        'Chegirma topilmadi.',
            'data':           None,
            'original_price': float(course.monthly_price),
        })

    return Response({'success': True, 'data': StudentDiscountSerializer(discount).data})


@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def cancel_discount(request, group_pk, student_group_pk):
    group         = get_object_or_404(Group, pk=group_pk)
    student_group = get_object_or_404(StudentGroup, pk=student_group_pk, group=group)
    student       = student_group.student
    course        = group.course
    today         = timezone.now().date()

    discount = StudentPricing.objects.filter(
        student=student, course=course,
        start_date__lte=today, end_date__gte=today,
    ).first()

    if not discount:
        return Response({'success': False, 'message': 'Aktiv chegirma topilmadi.'},
                        status=status.HTTP_404_NOT_FOUND)

    old_price        = str(discount.price_override)
    discount.end_date = today
    discount.save()

    # ── AuditLog ──
    _log('student', student.id, 'update',
         {'price_override': old_price},
         {'action': 'Chegirma bekor qilindi (cancel_discount)', 'end_date': str(today)},
         request.user)

    return Response({
        'success': True,
        'message': 'Chegirma bekor qilindi. DIQQAT: Balans avtomatik qaytarilmadi!',
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def discount_history(request, group_pk, student_group_pk):
    group         = get_object_or_404(Group, pk=group_pk)
    student_group = get_object_or_404(StudentGroup, pk=student_group_pk, group=group)
    student       = student_group.student
    course        = group.course

    discounts  = StudentPricing.objects.filter(student=student, course=course).order_by('-created_at')
    serializer = StudentDiscountSerializer(discounts, many=True)

    return Response({'success': True, 'count': discounts.count(), 'data': serializer.data})


# ════════════════════════════════════════════════════════════════
#  GURUH TARIXI - CLASS-BASED VIEWS
# ════════════════════════════════════════════════════════════════

from rest_framework.views import APIView
from audit.models import AuditLog, AuditEntityType
from .serializers import (
    GroupTeacherHistorySerializer,
    StudentGroupHistorySerializer,
    StudentLeaveHistorySerializer,
    StudentFreezeHistorySerializer,
)
from audit.serializers import AuditLogSerializer


class GroupHistoryView(APIView):
    """Guruh to'liq tarixi - barcha harakatlar"""

    def get(self, request, group_id):
        group = get_object_or_404(Group, id=group_id)

        teachers = GroupTeacher.objects.filter(
            group=group
        ).select_related('teacher').order_by('-created_at')

        students = StudentGroup.objects.filter(
            group=group
        ).select_related('student').order_by('-created_at')

        leaves = StudentGroupLeaves.objects.filter(
            group=group
        ).select_related('student', 'leave_reason', 'created_by').order_by('-created_at')

        freezes = StudentFreezes.objects.filter(
            group=group
        ).select_related('student', 'created_by').order_by('-created_at')

        audit_logs = AuditLog.objects.filter(
            entity_type=AuditEntityType.GROUP,
            entity_id=group.id,
        ).select_related('performed_by').order_by('-created_at')

        timeline = []

        for teacher in teachers:
            timeline.append({
                'type':   'teacher',
                'action': "O'qituvchi tayinlandi" if not teacher.end_date else "O'qituvchi almashtirildi",
                'date':   teacher.created_at,
                'data':   GroupTeacherHistorySerializer(teacher).data,
            })

        for student in students:
            timeline.append({
                'type':   'student_join',
                'action': "Talaba qo'shildi",
                'date':   student.created_at,
                'data':   StudentGroupHistorySerializer(student).data,
            })

        for leave in leaves:
            timeline.append({
                'type':   'student_leave',
                'action': 'Talaba ketdi',
                'date':   leave.created_at,
                'data':   StudentLeaveHistorySerializer(leave).data,
            })

        for freeze in freezes:
            timeline.append({
                'type':   'freeze',
                'action': 'Talaba muzlatildi',
                'date':   freeze.created_at,
                'data':   StudentFreezeHistorySerializer(freeze).data,
            })

        for log in audit_logs:
            timeline.append({
                'type':   'audit',
                'action': f"Guruh {log.get_action_display()}",
                'date':   log.created_at,
                'data':   AuditLogSerializer(log).data,
            })

        timeline.sort(key=lambda x: x['date'], reverse=True)

        return Response({
            'group': {
                'id':     str(group.id),
                'name':   group.name,
                'status': group.status,
            },
            'summary': {
                'total_teachers':        teachers.count(),
                'total_students_joined': students.count(),
                'total_students_left':   leaves.count(),
                'total_freezes':         freezes.count(),
                'total_changes':         len(timeline),
            },
            'timeline': timeline,
        })


class GroupTeacherHistoryView(APIView):
    """Guruh o'qituvchilari tarixi"""

    def get(self, request, group_id):
        group    = get_object_or_404(Group, id=group_id)
        teachers = GroupTeacher.objects.filter(
            group=group
        ).select_related('teacher').order_by('-created_at')

        return Response({
            'group_id':   str(group.id),
            'group_name': group.name,
            'teachers':   GroupTeacherHistorySerializer(teachers, many=True).data,
        })


class GroupStudentsHistoryView(APIView):
    """Guruh talabalari tarixi (qo'shilish va ketish)"""

    def get(self, request, group_id):
        group = get_object_or_404(Group, id=group_id)

        joined = StudentGroup.objects.filter(
            group=group
        ).select_related('student').order_by('-created_at')

        left = StudentGroupLeaves.objects.filter(
            group=group
        ).select_related('student', 'leave_reason', 'created_by').order_by('-created_at')

        return Response({
            'group_id':       str(group.id),
            'group_name':     group.name,
            'students_joined': StudentGroupHistorySerializer(joined, many=True).data,
            'students_left':  StudentLeaveHistorySerializer(left, many=True).data,
        })


class GroupFreezesHistoryView(APIView):
    """Guruh muzlatishlar tarixi"""

    def get(self, request, group_id):
        group   = get_object_or_404(Group, id=group_id)
        freezes = StudentFreezes.objects.filter(
            group=group
        ).select_related('student', 'created_by').order_by('-created_at')

        return Response({
            'group_id':   str(group.id),
            'group_name': group.name,
            'freezes':    StudentFreezeHistorySerializer(freezes, many=True).data,
        })


class GroupAuditLogView(APIView):
    """Guruh audit log - barcha o'zgarishlar"""

    def get(self, request, group_id):
        group = get_object_or_404(Group, id=group_id)
        logs  = AuditLog.objects.filter(
            entity_type=AuditEntityType.GROUP,
            entity_id=group.id,
        ).select_related('performed_by').order_by('-created_at')

        return Response({
            'group_id':   str(group.id),
            'group_name': group.name,
            'audit_logs': AuditLogSerializer(logs, many=True).data,
        })


# ════════════════════════════════════════════════════════════════
#  STUDENT TRANSFER / FREEZE / LEAVE - CLASS-BASED VIEWS
# ════════════════════════════════════════════════════════════════

from rest_framework.views import APIView
from django.db import transaction
from django.db.models import Q, Count, Prefetch
from datetime import date, timedelta
from django.http import HttpResponse


class StudentTransferView(APIView):
    """Talabani boshqa guruhga o'tkazish"""

    @transaction.atomic
    def post(self, request):
        from .serializers import StudentTransferSerializer, StudentGroupDetailSerializer
        serializer = StudentTransferSerializer(data=request.data)

        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        data           = serializer.validated_data
        student        = data['student']
        from_group     = data['from_group']
        to_group       = data['to_group']
        transfer_date  = data['transfer_date']
        recalc_balance = data['recalc_balance']

        # 1. Eski guruhdan chiqarish
        old_student_group = StudentGroup.objects.get(
            student=student, group=from_group, left_at__isnull=True
        )
        old_student_group.left_at = transfer_date
        old_student_group.save()

        # 2. Yangi guruhga qo'shish
        new_student_group = StudentGroup.objects.create(
            student=student, group=to_group,
            joined_at=transfer_date, end_date=to_group.end_date,
        )

        # 3. Balansni qayta hisoblash
        balance_info = None
        if recalc_balance:
            balance_info = self._recalculate_balance(student, from_group, to_group)

        # 4. Transaction yaratish
        if balance_info and balance_info.get('difference', 0) != 0:
            StudentTarnsactions.objects.create(
                student=student,
                student_group=new_student_group,
                group=to_group,
                transaction_type='correction',
                amount=balance_info['difference'],
                payment_type='transfer',
                transaction_date=timezone.now(),
                accepted_by=request.user,
                comment=f"Guruhdan o'tkazish: {from_group.name} → {to_group.name}",
            )

        # ── AuditLog ──
        _log('student', student.id, 'update',
             {'group_id': str(from_group.id), 'group': from_group.name},
             {
                 'action':         "Guruhga o'tkazildi",
                 'from_group_id':  str(from_group.id),
                 'from_group':     from_group.name,
                 'to_group_id':    str(to_group.id),
                 'to_group':       to_group.name,
                 'transfer_date':  str(transfer_date),
                 'balance_info':   balance_info,
             }, request.user)

        return Response({
            'message':             "Talaba muvaffaqiyatli o'tkazildi",
            'old_group':           {'id': str(from_group.id), 'name': from_group.name},
            'new_group':           {'id': str(to_group.id), 'name': to_group.name},
            'student':             {'id': str(student.id), 'full_name': student.full_name},
            'balance_recalculated': recalc_balance,
            'balance_info':        balance_info,
        }, status=status.HTTP_200_OK)

    def _recalculate_balance(self, student, from_group, to_group):
        try:
            balance     = StudentBalances.objects.get(student=student)
            old_balance = balance.balance
            old_price   = from_group.course.monthly_price
            new_price   = to_group.course.monthly_price
            difference  = new_price - old_price
            balance.balance += difference
            balance.save()
            return {
                'old_balance': str(old_balance),
                'new_balance': str(balance.balance),
                'difference':  str(difference),
                'old_price':   str(old_price),
                'new_price':   str(new_price),
            }
        except StudentBalances.DoesNotExist:
            return None


class StudentFreezeView(APIView):
    """Talabani muzlatish"""

    @transaction.atomic
    def post(self, request):
        from .serializers import StudentFreezeSerializer, StudentFreezeDetailSerializer
        serializer = StudentFreezeSerializer(data=request.data, context={'request': request})

        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        freeze = serializer.save()

        if freeze.recalc_balance:
            self._recalculate_freeze_balance(freeze)

        # ── AuditLog ──
        _log('student', freeze.student_id, 'update', None, {
            'action':            'Muzlatildi (ClassView)',
            'group_id':          str(freeze.group_id),
            'freeze_start_date': str(freeze.freeze_start_date),
            'freeze_end_date':   str(freeze.freeze_end_date),
            'reason':            freeze.reason or '',
        }, request.user)

        return Response({
            'message': "Talaba muvaffaqiyatli muzlatildi",
            'freeze':  StudentFreezeDetailSerializer(freeze).data,
        }, status=status.HTTP_201_CREATED)

    def get(self, request):
        from .serializers import StudentFreezeDetailSerializer
        student_id = request.query_params.get('student_id')
        group_id   = request.query_params.get('group_id')

        freezes = StudentFreezes.objects.all().select_related('student', 'group', 'created_by')
        if student_id:
            freezes = freezes.filter(student_id=student_id)
        if group_id:
            freezes = freezes.filter(group_id=group_id)
        freezes = freezes.order_by('-created_at')

        return Response({
            'count':   freezes.count(),
            'freezes': StudentFreezeDetailSerializer(freezes, many=True).data,
        })

    def _recalculate_freeze_balance(self, freeze):
        try:
            balance      = StudentBalances.objects.get(student=freeze.student)
            freeze_days  = (freeze.freeze_end_date - freeze.freeze_start_date).days
            daily_price  = freeze.group.course.monthly_price / 30
            refund_amount = daily_price * freeze_days

            balance.balance += refund_amount
            balance.save()

            StudentTarnsactions.objects.create(
                student=freeze.student,
                group=freeze.group,
                transaction_type='refund',
                amount=refund_amount,
                payment_type='correction',
                transaction_date=timezone.now(),
                comment=f"Muzlatish uchun qaytarildi: {freeze_days} kun",
            )
        except StudentBalances.DoesNotExist:
            pass


class StudentLeaveView(APIView):
    """Talabani guruhdan chiqarish"""

    @transaction.atomic
    def post(self, request):
        from .serializers import StudentLeaveSerializer
        serializer = StudentLeaveSerializer(data=request.data)

        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        data          = serializer.validated_data
        student       = data['student']
        group         = data['group']
        student_group = data['student_group']

        # 1. StudentGroup yangilash
        student_group.left_at = data['leave_date']
        student_group.save()

        # 2. Leave yaratish
        leave_reason = None
        if data.get('leave_reason_id'):
            try:
                leave_reason = LeaveReason.objects.get(id=data['leave_reason_id'])
            except LeaveReason.DoesNotExist:
                pass

        leave = StudentGroupLeaves.objects.create(
            student=student,
            group=group,
            student_group=student_group,
            leave_date=data['leave_date'],
            leave_reason=leave_reason,
            comment=data.get('comment', ''),
            recalc_balance=data['recalc_balance'],
            refound_amount=data['refound_amount'],
            created_by=request.user,
        )

        # 3. Pul qaytarish
        if data['recalc_balance'] and data['refound_amount'] > 0:
            self._process_refund(student, group, data['refound_amount'])

        # ── AuditLog ──
        _log('student', student.id, 'update',
             {'group_id': str(group.id), 'group': group.name},
             {
                 'action':         "Guruhdan chiqarildi (ClassView)",
                 'group_id':       str(group.id),
                 'group':          group.name,
                 'leave_date':     str(leave.leave_date),
                 'leave_reason':   leave_reason.name if leave_reason else None,
                 'refound_amount': str(leave.refound_amount),
             }, request.user)

        return Response({
            'message':       "Talaba guruhdan muvaffaqiyatli chiqarildi",
            'student':       {'id': str(student.id), 'full_name': student.full_name},
            'group':         {'id': str(group.id), 'name': group.name},
            'leave_date':    leave.leave_date,
            'refound_amount': str(leave.refound_amount),
        }, status=status.HTTP_200_OK)

    def _process_refund(self, student, group, amount):
        try:
            balance = StudentBalances.objects.get(student=student)
            balance.balance += amount
            balance.save()
            StudentTarnsactions.objects.create(
                student=student, group=group,
                transaction_type='refund', amount=amount,
                payment_type='correction',
                transaction_date=timezone.now(),
                comment="Guruhdan chiqish - pul qaytarish",
            )
        except StudentBalances.DoesNotExist:
            pass


class StudentActiveGroupsView(APIView):
    """Talabaning faol guruhlari"""

    def get(self, request, student_id):
        from .serializers import StudentGroupDetailSerializer
        groups = StudentGroup.objects.filter(
            student_id=student_id, left_at__isnull=True
        ).select_related('group', 'student')

        return Response({
            'count':  groups.count(),
            'groups': StudentGroupDetailSerializer(groups, many=True).data,
        })


class StudentFreezesListView(APIView):
    """Talabaning muzlatishlari"""

    def get(self, request, student_id):
        from .serializers import StudentFreezeDetailSerializer
        freezes = StudentFreezes.objects.filter(
            student_id=student_id
        ).select_related('group', 'student', 'created_by').order_by('-created_at')

        return Response({
            'count':   freezes.count(),
            'freezes': StudentFreezeDetailSerializer(freezes, many=True).data,
        })


class BalanceRecalculationView(APIView):
    """Balansni qo'lda qayta hisoblash"""

    @transaction.atomic
    def post(self, request):
        from .serializers import BalanceRecalculationSerializer
        serializer = BalanceRecalculationSerializer(data=request.data)

        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        data        = serializer.validated_data
        student     = data['student']
        recalc_type = data['recalc_type']

        result = self._calculate_balance(student, recalc_type, data)

        # ── AuditLog ──
        _log('student', student.id, 'update',
             {'balance': result.get('old_balance')},
             {
                 'action':      'Balans qayta hisoblandi',
                 'recalc_type': recalc_type,
                 'new_balance': result.get('new_balance'),
             }, request.user)

        return Response({
            'message':      "Balans qayta hisoblandi",
            'student':      {'id': str(student.id), 'full_name': student.full_name},
            'balance_info': result,
        })

    def _calculate_balance(self, student, recalc_type, data):
        try:
            balance     = StudentBalances.objects.get(student=student)
            old_balance = balance.balance
            # Hisoblash logikasi (recalc_type ga qarab kengaytirish mumkin)
            balance.save()
            return {
                'old_balance': str(old_balance),
                'new_balance': str(balance.balance),
                'recalc_type': recalc_type,
            }
        except StudentBalances.DoesNotExist:
            return {'error': 'Balans topilmadi'}


# ════════════════════════════════════════════════════════════════
#  GURUHLAR RO'YXATI - CLASS-BASED VIEWS
# ════════════════════════════════════════════════════════════════

class GroupsListView(APIView):
    """Guruhlar ro'yxati - filtrlash, qidiruv, pagination"""

    def get(self, request):
        from .serializers import GroupFilterSerializer, GroupListSerializer
        filter_serializer = GroupFilterSerializer(data=request.query_params)
        filter_serializer.is_valid()
        filters = filter_serializer.validated_data

        queryset = Group.objects.all().select_related('course', 'room').prefetch_related(
            Prefetch('group', queryset=GroupTeacher.objects.select_related('teacher')),
            Prefetch('Sgroup_group', queryset=StudentGroup.objects.select_related('student')),
        )

        queryset = self._apply_filters(queryset, filters)

        sort_by   = request.query_params.get('sort_by', '-created_at')
        queryset  = queryset.order_by(sort_by)

        page       = int(request.query_params.get('page', 1))
        page_size  = int(request.query_params.get('page_size', 20))
        start      = (page - 1) * page_size
        total_count = queryset.count()
        groups     = queryset[start: start + page_size]

        return Response({
            'count':       total_count,
            'page':        page,
            'page_size':   page_size,
            'total_pages': (total_count + page_size - 1) // page_size,
            'results':     GroupListSerializer(groups, many=True).data,
        })

    def _apply_filters(self, queryset, filters):
        if 'status' in filters:
            queryset = queryset.filter(status=filters['status'])
        if 'course_id' in filters:
            queryset = queryset.filter(course_id=filters['course_id'])
        if 'teacher_id' in filters:
            queryset = queryset.filter(
                group__teacher_id=filters['teacher_id'],
                group__end_date__isnull=True,
            )
        if 'room_id' in filters:
            queryset = queryset.filter(room_id=filters['room_id'])
        if 'search' in filters:
            queryset = queryset.filter(
                Q(name__icontains=filters['search']) |
                Q(course__name__icontains=filters['search'])
            )
        if 'start_date_from' in filters:
            queryset = queryset.filter(start_date__gte=filters['start_date_from'])
        if 'start_date_to' in filters:
            queryset = queryset.filter(start_date__lte=filters['start_date_to'])
        if 'has_students' in filters:
            ann = queryset.annotate(
                student_count=Count('Sgroup_group', filter=Q(Sgroup_group__left_at__isnull=True))
            )
            queryset = ann.filter(student_count__gt=0) if filters['has_students'] else ann.filter(student_count=0)
        if 'days_until_end_lt' in filters:
            target = date.today() + timedelta(days=filters['days_until_end_lt'])
            queryset = queryset.filter(end_date__lte=target, end_date__gte=date.today())
        return queryset


class GroupCreateView(APIView):
    """Guruh yaratish"""

    def post(self, request):
        from .serializers import GroupCreateSerializer, GroupDetailSerializer
        serializer = GroupCreateSerializer(data=request.data)

        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        group = serializer.save()

        # ── AuditLog ──
        _log('group', group.id, 'create', None, {
            'name':       group.name,
            'course_id':  str(group.course_id),
            'status':     group.status,
            'start_date': str(group.start_date),
            'end_date':   str(group.end_date),
        }, request.user)

        return Response({
            'message': 'Guruh muvaffaqiyatli yaratildi',
            'group':   GroupDetailSerializer(group).data,
        }, status=status.HTTP_201_CREATED)


class GroupDetailView(APIView):
    """Guruh batafsil ma'lumot, tahrirlash, o'chirish"""

    def get(self, request, group_id):
        from .serializers import GroupDetailSerializer
        group = get_object_or_404(Group, id=group_id)
        return Response(GroupDetailSerializer(group).data)

    def put(self, request, group_id):
        from .serializers import GroupDetailSerializer, GroupUpdateSerializer
        group    = get_object_or_404(Group, id=group_id)
        old_data = {'name': group.name, 'status': group.status, 'end_date': str(group.end_date)}

        serializer = GroupUpdateSerializer(group, data=request.data, partial=True)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        serializer.save()

        # ── AuditLog ──
        _log('group', group.id, 'update', old_data, {
            'name':     group.name,
            'status':   group.status,
            'end_date': str(group.end_date),
        }, request.user)

        return Response({
            'message': 'Guruh muvaffaqiyatli yangilandi',
            'group':   GroupDetailSerializer(group).data,
        })

    def delete(self, request, group_id):
        group = get_object_or_404(Group, id=group_id)

        has_students = StudentGroup.objects.filter(group=group, left_at__isnull=True).exists()
        if has_students:
            return Response({'error': 'Guruhda faol talabalar bor. Avval ularni chiqaring.'},
                            status=status.HTTP_400_BAD_REQUEST)

        # ── AuditLog ──
        _log('group', group.id, 'delete',
             {'name': group.name, 'status': group.status},
             None, request.user)

        group_name = group.name
        group.delete()
        return Response({'message': f"{group_name} guruhi o'chirildi"})


class GroupArchiveView(APIView):
    """Guruhni arxivlash"""

    def post(self, request, group_id):
        from .serializers import GroupArchiveSerializer
        group      = get_object_or_404(Group, id=group_id)
        serializer = GroupArchiveSerializer(data=request.data)

        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        old_status   = group.status
        group.status = 'archived'
        group.save()

        # ── AuditLog ──
        _log('group', group.id, 'update',
             {'status': old_status},
             {'status': 'archived', 'reason': serializer.validated_data.get('reason', '')},
             request.user)

        return Response({
            'message': 'Guruh arxivlandi',
            'group':   {'id': str(group.id), 'name': group.name, 'status': group.status},
        })


class GroupUnarchiveView(APIView):
    """Guruhni arxivdan qaytarish"""

    def post(self, request, group_id):
        group = get_object_or_404(Group, id=group_id)

        if group.status != 'archived':
            return Response({'error': 'Guruh arxivda emas'}, status=status.HTTP_400_BAD_REQUEST)

        group.status = 'active'
        group.save()

        # ── AuditLog ──
        _log('group', group.id, 'update',
             {'status': 'archived'},
             {'status': 'active'},
             request.user)

        return Response({
            'message': 'Guruh qayta faollashtirildi',
            'group':   {'id': str(group.id), 'name': group.name, 'status': group.status},
        })


class GroupExportExcelView(APIView):
    """Guruhlarni Excel ga export qilish"""

    def get(self, request):
        from .serializers import GroupFilterSerializer
        filter_serializer = GroupFilterSerializer(data=request.query_params)
        filter_serializer.is_valid()
        filters  = filter_serializer.validated_data
        queryset = Group.objects.all().select_related('course', 'room')
        queryset = GroupsListView()._apply_filters(queryset, filters)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Guruhlar"

        headers = ['Guruh nomi', 'Kurs', 'Status', 'Xona',
                   'Boshlanish', 'Tugash', 'Talabalar soni', "O'qituvchilar", 'Yaratilgan']
        ws.append(headers)

        for group in queryset:
            students_count = StudentGroup.objects.filter(group=group, left_at__isnull=True).count()
            teachers       = GroupTeacher.objects.filter(group=group, end_date__isnull=True).select_related('teacher')
            teacher_names  = ', '.join([f"{t.teacher.user.full_name}" for t in teachers])

            ws.append([
                group.name,
                group.course.name if group.course else '',
                group.get_status_display(),
                group.room.name if group.room else '',
                group.start_date.strftime('%d.%m.%Y') if group.start_date else '',
                group.end_date.strftime('%d.%m.%Y') if group.end_date else '',
                students_count,
                teacher_names,
                group.created_at.strftime('%d.%m.%Y %H:%M'),
            ])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=groups.xlsx'
        wb.save(response)
        return response


class GroupStatisticsView(APIView):
    """Guruhlar statistikasi"""

    def get(self, request):
        return Response({
            'total_groups':   Group.objects.count(),
            'active_groups':  Group.objects.filter(status='active').count(),
            'expired_groups': Group.objects.filter(status='expired').count(),
            'archived_groups': Group.objects.filter(status='archived').count(),
            'ending_soon':    Group.objects.filter(
                status='active',
                end_date__lte=date.today() + timedelta(days=7),
                end_date__gte=date.today(),
            ).count(),
            'total_students': StudentGroup.objects.filter(
                left_at__isnull=True
            ).values('student').distinct().count(),
        })


class GroupBulkActionView(APIView):
    """Bir nechta guruhga bir vaqtda amallar bajarish"""

    def post(self, request):
        action    = request.data.get('action')
        group_ids = request.data.get('group_ids', [])

        if not action or not group_ids:
            return Response({'error': 'action va group_ids majburiy'}, status=status.HTTP_400_BAD_REQUEST)

        groups = Group.objects.filter(id__in=group_ids)

        if action == 'archive':
            count = groups.count()
            groups.update(status='archived')

            # ── AuditLog (har bir guruh uchun) ──
            for group in Group.objects.filter(id__in=group_ids):
                _log('group', group.id, 'update', {'status': 'active'},
                     {'status': 'archived', 'action': 'bulk_archive'}, request.user)

            message = f"{count} ta guruh arxivlandi"

        elif action == 'unarchive':
            count = groups.count()
            groups.update(status='active')

            for group in Group.objects.filter(id__in=group_ids):
                _log('group', group.id, 'update', {'status': 'archived'},
                     {'status': 'active', 'action': 'bulk_unarchive'}, request.user)

            message = f"{count} ta guruh qayta faollashtirildi"

        elif action == 'delete':
            count = 0
            for group in groups:
                has_students = StudentGroup.objects.filter(group=group, left_at__isnull=True).exists()
                if not has_students:
                    _log('group', group.id, 'delete',
                         {'name': group.name}, {'action': 'bulk_delete'}, request.user)
                    group.delete()
                    count += 1
            message = f"{count} ta guruh o'chirildi"

        else:
            return Response({'error': "Noto'g'ri action"}, status=status.HTTP_400_BAD_REQUEST)

        return Response({'message': message})


# ════════════════════════════════════════════════════════════════
#  EXPORT VIEWS
# ════════════════════════════════════════════════════════════════

class GroupsExportView(APIView):
    """Guruhlarni Excel ga export qilish"""

    def post(self, request):
        filter_serializer  = GroupExportFilterSerializer(data=request.data.get('filters', {}))
        columns_serializer = GroupExportColumnsSerializer(data=request.data.get('columns', {}))

        if not filter_serializer.is_valid():
            return Response(filter_serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        if not columns_serializer.is_valid():
            return Response(columns_serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        filters     = filter_serializer.validated_data
        all_columns = (
            columns_serializer.validated_data.get('default_columns', []) +
            columns_serializer.validated_data.get('additional_columns', [])
        )

        queryset   = self._get_filtered_groups(filters)
        file_name  = self._generate_filename('Groups')
        excel_file = self._create_groups_excel(queryset, all_columns)

        # Export audit log (mavjud helper orqali)
        create_export_audit_log(
            user=request.user,
            export_type='groups',
            file_name=file_name,
            total_records=queryset.count(),
            columns=all_columns,
            filters=filters,
        )

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename={file_name}'
        excel_file.save(response)
        return response

    def _get_filtered_groups(self, filters):
        queryset = Group.objects.all().select_related('course', 'room')
        if filters.get('group_ids'):
            queryset = queryset.filter(id__in=filters['group_ids'])
        if filters.get('status'):
            queryset = queryset.filter(status=filters['status'])
        if filters.get('course_id'):
            queryset = queryset.filter(course_id=filters['course_id'])
        if filters.get('teacher_id'):
            queryset = queryset.filter(
                group__teacher_id=filters['teacher_id'], group__end_date__isnull=True
            )
        if filters.get('room_id'):
            queryset = queryset.filter(room_id=filters['room_id'])
        if filters.get('start_date_from'):
            queryset = queryset.filter(start_date__gte=filters['start_date_from'])
        if filters.get('start_date_to'):
            queryset = queryset.filter(start_date__lte=filters['start_date_to'])
        return queryset.order_by('name')

    def _create_groups_excel(self, groups, columns):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Guruhlar"

        serializer = GroupExportDataSerializer(context={'columns': columns})
        headers    = list(serializer.to_representation(groups.first()).keys()) if groups.exists() else ["Ma'lumot yo'q"]

        header_font      = Font(bold=True, color="FFFFFF", size=11)
        header_fill      = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border           = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'),  bottom=Side(style='thin'),
        )

        for col_num, header in enumerate(headers, 1):
            cell           = ws.cell(row=1, column=col_num, value=header)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = header_alignment
            cell.border    = border

        for row_num, group in enumerate(groups, 2):
            data = serializer.to_representation(group)
            for col_num, value in enumerate(data.values(), 1):
                cell           = ws.cell(row=row_num, column=col_num, value=value)
                cell.border    = border
                cell.alignment = Alignment(vertical="center")

        for column in ws.columns:
            max_length    = max((len(str(c.value or '')) for c in column), default=0)
            ws.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)

        return wb

    def _generate_filename(self, prefix):
        return f"{prefix}Export_{datetime.now().strftime('%Y-%m-%d_%H%M')}.xlsx"


class StudentsExportView(APIView):
    """Talabalarni Excel ga export qilish"""

    def post(self, request):
        filter_serializer  = StudentExportFilterSerializer(data=request.data.get('filters', {}))
        columns_serializer = StudentExportColumnsSerializer(data=request.data.get('columns', {}))

        if not filter_serializer.is_valid():
            return Response(filter_serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        if not columns_serializer.is_valid():
            return Response(columns_serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        filters     = filter_serializer.validated_data
        all_columns = (
            columns_serializer.validated_data.get('default_columns', []) +
            columns_serializer.validated_data.get('additional_columns', [])
        )

        queryset   = self._get_filtered_students(filters)
        group_name = queryset.first().group.name if queryset.exists() else 'Students'
        file_name  = self._generate_filename(group_name)
        excel_file = self._create_students_excel(queryset, all_columns)

        create_export_audit_log(
            user=request.user,
            export_type='students',
            file_name=file_name,
            total_records=queryset.count(),
            columns=all_columns,
            filters=filters,
        )

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename={file_name}'
        excel_file.save(response)
        return response

    def _get_filtered_students(self, filters):
        group_id      = filters['group_id']
        student_types = filters.get('student_types', ['active'])

        queryset = StudentGroup.objects.filter(group_id=group_id).select_related('student', 'group')
        q_filter = Q()

        if 'active' in student_types:
            q_filter |= Q(left_at__isnull=True)
        if 'left' in student_types:
            q_filter |= Q(left_at__isnull=False)
        if 'frozen' in student_types:
            frozen_ids = StudentFreezes.objects.filter(
                group_id=group_id,
                freeze_end_date__gte=datetime.now().date(),
            ).values_list('student_id', flat=True)
            q_filter |= Q(student_id__in=frozen_ids)

        return queryset.filter(q_filter).distinct().order_by('student__full_name')

    def _create_students_excel(self, students, columns):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Talabalar"

        serializer = StudentExportDataSerializer(context={'columns': columns})
        headers    = list(serializer.to_representation(students.first()).keys()) if students.exists() else ["Ma'lumot yo'q"]

        header_font      = Font(bold=True, color="FFFFFF", size=11)
        header_fill      = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border           = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'),  bottom=Side(style='thin'),
        )

        for col_num, header in enumerate(headers, 1):
            cell           = ws.cell(row=1, column=col_num, value=header)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = header_alignment
            cell.border    = border

        for row_num, student_group in enumerate(students, 2):
            data = serializer.to_representation(student_group)
            for col_num, value in enumerate(data.values(), 1):
                cell           = ws.cell(row=row_num, column=col_num, value=value)
                cell.border    = border
                cell.alignment = Alignment(vertical="center")

        for column in ws.columns:
            max_length = max((len(str(c.value or '')) for c in column), default=0)
            ws.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)

        return wb

    def _generate_filename(self, group_name):
        safe_name = group_name.replace(' ', '_').replace('/', '-')
        return f"GuruhStudent_{safe_name}_{datetime.now().strftime('%Y-%m-%d_%H%M')}.xlsx"


# ════════════════════════════════════════════════════════════════
#  EXPORT HISTORY / STATISTICS
# ════════════════════════════════════════════════════════════════

class ExportHistoryView(APIView):
    """Export tarixi - Audit log dan"""

    def get(self, request):
        logs = AuditLog.objects.filter(
            action=AuditAction.CREATE,
            new_action__isnull=False,
        ).select_related('performed_by').order_by('-created_at')

        export_type = request.query_params.get('export_type')
        if export_type:
            logs = logs.filter(new_action__export_type=export_type)

        page      = int(request.query_params.get('page', 1))
        page_size = int(request.query_params.get('page_size', 20))
        start     = (page - 1) * page_size
        total     = logs.count()
        logs      = logs[start: start + page_size]

        return Response({
            'count':       total,
            'page':        page,
            'page_size':   page_size,
            'total_pages': (total + page_size - 1) // page_size,
            'results':     ExportHistorySerializer(logs, many=True).data,
        })


class ExportStatisticsView(APIView):
    """Export statistikasi"""

    def get(self, request):
        total_exports    = AuditLog.objects.filter(action=AuditAction.CREATE, new_action__isnull=False).count()
        groups_exports   = AuditLog.objects.filter(action=AuditAction.CREATE, new_action__export_type='groups').count()
        students_exports = AuditLog.objects.filter(action=AuditAction.CREATE, new_action__export_type='students').count()

        last_export      = AuditLog.objects.filter(
            action=AuditAction.CREATE, new_action__isnull=False
        ).order_by('-created_at').first()

        last_export_data = None
        if last_export:
            last_export_data = {
                'type':         last_export.new_action.get('export_type'),
                'file_name':    last_export.new_action.get('file_name'),
                'date':         last_export.created_at,
                'performed_by': last_export.performed_by.user.full_name if last_export.performed_by else None,
            }

        return Response({
            'total_exports':    total_exports,
            'groups_exports':   groups_exports,
            'students_exports': students_exports,
            'last_export':      last_export_data,
        })


# ════════════════════════════════════════════════════════════════
#  IMTIHONLAR - CLASS-BASED VIEWS
# ════════════════════════════════════════════════════════════════

class ExamsListView(APIView):
    """Imtihonlar ro'yxati"""

    def get(self, request):
        group_id = request.query_params.get('group_id')
        exams    = Exams.objects.all().select_related('group', 'created_by').order_by('-exam_date')

        if group_id:
            exams = exams.filter(group_id=group_id)

        page      = int(request.query_params.get('page', 1))
        page_size = int(request.query_params.get('page_size', 20))
        start     = (page - 1) * page_size
        total     = exams.count()
        exams     = exams[start: start + page_size]

        return Response({
            'count':       total,
            'page':        page,
            'page_size':   page_size,
            'total_pages': (total + page_size - 1) // page_size,
            'results':     ExamListSerializer(exams, many=True).data,
        })


class ExamCreateView(APIView):
    """Yangi imtihon yaratish"""

    def post(self, request):
        serializer = ExamCreateSerializer(data=request.data, context={'request': request})

        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        exam = serializer.save()

        # ── AuditLog ──
        _log('other', exam.id, 'create', None, {
            'action':    'Imtihon yaratildi (ExamCreateView)',
            'title':     exam.title,
            'group_id':  str(exam.group_id),
            'exam_date': str(exam.exam_date),
        }, request.user)

        return Response({
            'message': 'Imtihon muvaffaqiyatli yaratildi',
            'exam':    ExamDetailSerializer(exam).data,
        }, status=status.HTTP_201_CREATED)


class ExamDetailView(APIView):
    """Imtihon batafsil ma'lumot, tahrirlash, o'chirish"""

    def get(self, request, exam_id):
        exam = get_object_or_404(Exams, id=exam_id)
        return Response(ExamDetailSerializer(exam).data)

    def put(self, request, exam_id):
        exam       = get_object_or_404(Exams, id=exam_id)
        old_data   = {'title': exam.title, 'exam_date': str(exam.exam_date)}
        serializer = ExamUpdateSerializer(exam, data=request.data, partial=True)

        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        serializer.save()

        # ── AuditLog ──
        _log('other', exam.id, 'update', old_data, {
            'title':     exam.title,
            'exam_date': str(exam.exam_date),
        }, request.user)

        return Response({
            'message': 'Imtihon muvaffaqiyatli yangilandi',
            'exam':    ExamDetailSerializer(exam).data,
        })

    def delete(self, request, exam_id):
        exam        = get_object_or_404(Exams, id=exam_id)
        has_results = ExamResults.objects.filter(exam=exam).exists()

        if has_results:
            return Response(
                {'error': "Imtihonda natijalar mavjud. Avval natijalarni o'chiring."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # ── AuditLog ──
        _log('other', exam.id, 'delete',
             {'title': exam.title, 'group_id': str(exam.group_id)},
             None, request.user)

        exam_title = exam.title
        exam.delete()
        return Response({'message': f'"{exam_title}" imtihoni o\'chirildi'})


class ExamGradingView(APIView):
    """Imtihon natijalarini kiritish/yangilash"""

    def post(self, request):
        if 'student_id' in request.data:
            return self._grade_single_student(request)
        elif 'grades' in request.data:
            return self._grade_multiple_students(request)
        return Response({'error': 'student_id yoki grades parametri kerak'}, status=status.HTTP_400_BAD_REQUEST)

    def _grade_single_student(self, request):
        serializer = ExamResultSerializer(data=request.data, context={'request': request})

        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        result = serializer.save()

        # ── AuditLog ──
        _log('other', result.exam_id, 'update', None, {
            'action':     'Ball kiritildi',
            'student_id': str(result.student_id),
            'score':      str(result.score),
        }, request.user)

        return Response({
            'message': 'Ball muvaffaqiyatli kiritildi',
            'result':  {
                'id':     str(result.id),
                'score':  result.score,
                'status': 'passed' if result.score >= result.exam.min_score else 'failed',
            },
        })

    def _grade_multiple_students(self, request):
        grades  = request.data.get('grades', [])
        exam_id = request.data.get('exam_id')

        if not exam_id:
            return Response({'error': 'exam_id parametri kerak'}, status=status.HTTP_400_BAD_REQUEST)

        results = []
        errors  = []

        for grade in grades:
            grade['exam_id'] = exam_id
            serializer = ExamResultSerializer(data=grade, context={'request': request})

            if serializer.is_valid():
                result = serializer.save()
                results.append({
                    'student_id': grade['student_id'],
                    'score':      result.score,
                    'status':     'success',
                })
            else:
                errors.append({'student_id': grade.get('student_id'), 'errors': serializer.errors})

        # ── AuditLog (bulk) ──
        if results:
            _log('other', exam_id, 'update', None, {
                'action':        'Bulk ball kiritildi',
                'saved_count':   len(results),
                'error_count':   len(errors),
            }, request.user)

        return Response({
            'message': f'{len(results)} ta natija kiritildi',
            'results': results,
            'errors':  errors if errors else None,
        })


class ExamResultsView(APIView):
    """Imtihon natijalari ro'yxati"""

    def get(self, request, exam_id):
        exam = get_object_or_404(Exams, id=exam_id)

        eligible_students = StudentGroup.objects.filter(
            group=exam.group, left_at__isnull=True
        ).select_related('student')

        results = []
        for sg in eligible_students:
            student     = sg.student
            exam_result = ExamResults.objects.filter(exam=exam, student=student).first()

            if exam_result:
                results.append({
                    'student_id':   str(student.id),
                    'full_name':    student.full_name,
                    'phone_number': student.phone_number,
                    'score':        exam_result.score,
                    'comment':      exam_result.comment,
                    'status':       'passed' if exam_result.score >= exam.min_score else 'failed',
                    'status_color': 'green' if exam_result.score >= exam.min_score else 'red',
                    'graded':       True,
                    'graded_at':    exam_result.created_at,
                })
            else:
                results.append({
                    'student_id':   str(student.id),
                    'full_name':    student.full_name,
                    'phone_number': student.phone_number,
                    'score':        None,
                    'comment':      '',
                    'status':       'not_graded',
                    'status_color': 'gray',
                    'graded':       False,
                    'graded_at':    None,
                })

        return Response({
            'exam': {
                'id':        str(exam.id),
                'title':     exam.title,
                'exam_date': exam.exam_date,
                'min_score': exam.min_score,
                'max_score': exam.max_score,
            },
            'results': results,
        })


class ExamStatisticsView(APIView):
    """Imtihon statistikasi"""

    def get(self, request, exam_id):
        exam    = get_object_or_404(Exams, id=exam_id)
        results = ExamResults.objects.filter(exam=exam)

        total_students = StudentGroup.objects.filter(group=exam.group, left_at__isnull=True).count()
        graded         = results.count()
        not_graded     = total_students - graded
        passed         = results.filter(score__gte=exam.min_score).count()
        failed         = results.filter(score__lt=exam.min_score).count()
        avg_score      = results.aggregate(avg=Avg('score'))['avg'] or 0

        highest_score = lowest_score = None
        if results.exists():
            max_r = results.order_by('-score').first()
            min_r = results.order_by('score').first()
            highest_score = {'score': max_r.score, 'student': max_r.student.full_name}
            lowest_score  = {'score': min_r.score, 'student': min_r.student.full_name}

        step               = exam.max_score / 5
        score_distribution = []
        for i in range(5):
            min_r = i * step
            max_r = (i + 1) * step
            score_distribution.append({
                'range': f'{int(min_r)}-{int(max_r)}',
                'count': results.filter(score__gte=min_r, score__lt=max_r).count(),
            })

        return Response({
            'total_students':     total_students,
            'graded':             graded,
            'not_graded':         not_graded,
            'passed':             passed,
            'failed':             failed,
            'pass_rate':          round((passed / graded * 100), 1) if graded > 0 else 0,
            'average_score':      round(float(avg_score), 2),
            'highest_score':      highest_score,
            'lowest_score':       lowest_score,
            'score_distribution': score_distribution,
        })


class ExamParticipantsSettingsView(APIView):
    """Imtihon qatnashuvchilar sozlamalari (CEO uchun)"""

    def get(self, request):
        return Response({
            'include_active': True,
            'include_trial':  False,
            'include_frozen': False,
            'include_left':   False,
        })

    def post(self, request):
        if not hasattr(request.user, 'role') or request.user.role != 'ceo':
            return Response({'error': 'Faqat CEO sozlamalarga kirish huquqiga ega'},
                            status=status.HTTP_403_FORBIDDEN)

        serializer = ExamParticipantsSettingsSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        return Response({'message': 'Sozlamalar yangilandi', 'settings': serializer.validated_data})


# ════════════════════════════════════════════════════════════════
#  ONLAYN DARSLAR - CLASS-BASED VIEWS
# ════════════════════════════════════════════════════════════════

class OnlineLessonsListView(APIView):
    """Onlayn darslar ro'yxati"""

    def get(self, request):
        group_id     = request.query_params.get('group_id')
        is_published = request.query_params.get('is_published')

        lessons = OnlineLesson.objects.all().select_related('group', 'created_by').order_by('order', '-lesson_date')

        if group_id:
            lessons = lessons.filter(group_id=group_id)
        if is_published is not None:
            lessons = lessons.filter(is_published=is_published.lower() == 'true')

        page      = int(request.query_params.get('page', 1))
        page_size = int(request.query_params.get('page_size', 20))
        start     = (page - 1) * page_size
        total     = lessons.count()
        lessons   = lessons[start: start + page_size]

        return Response({
            'count':       total,
            'page':        page,
            'page_size':   page_size,
            'total_pages': (total + page_size - 1) // page_size,
            'results':     OnlineLessonListSerializer(lessons, many=True).data,
        })


class SetLessonTopicView(APIView):
    """Dars mavzusini belgilash"""

    def post(self, request):
        serializer = LessonTopicSerializer(data=request.data, context={'request': request})

        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        lesson = serializer.save()

        # ── AuditLog ──
        _log('other', lesson.id, 'create', None, {
            'action':   'Dars mavzusi belgilandi',
            'title':    lesson.title,
            'group_id': str(lesson.group_id),
            'date':     str(lesson.lesson_date),
        }, request.user)

        return Response({
            'message': 'Mavzu belgilandi va onlayn dars yaratildi',
            'lesson':  OnlineLessonDetailSerializer(lesson).data,
        }, status=status.HTTP_201_CREATED)


class UpdateLessonTopicView(APIView):
    """Dars mavzusini o'zgartirish"""

    def put(self, request):
        lesson_date     = request.data.get('lesson_date')
        group_id        = request.data.get('group_id')
        new_title       = request.data.get('title')
        new_description = request.data.get('description', '')

        if not all([lesson_date, group_id, new_title]):
            return Response({'error': 'lesson_date, group_id va title kerak'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            lesson    = OnlineLesson.objects.get(group_id=group_id, lesson_date=lesson_date)
            old_title = lesson.title

            lesson.title       = new_title
            lesson.description = new_description
            lesson.save()

            # ── AuditLog ──
            _log('other', lesson.id, 'update',
                 {'title': old_title},
                 {'title': new_title, 'action': 'Dars mavzusi yangilandi'},
                 request.user)

            return Response({'message': 'Mavzu yangilandi', 'lesson': OnlineLessonDetailSerializer(lesson).data})

        except OnlineLesson.DoesNotExist:
            return Response({'error': 'Bu sana uchun onlayn dars topilmadi'}, status=status.HTTP_404_NOT_FOUND)


class GroupLessonsCalendarView(APIView):
    """Guruh dars kalendari"""

    def get(self, request):
        serializer = GroupLessonsCalendarSerializer(data=request.query_params)

        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        data       = serializer.validated_data
        group      = data['group']
        start_date = data['start_date']
        end_date   = data['end_date']
        schedule   = LessonSchedule.objects.filter(group=group)

        calendar_data = []
        current_date  = start_date

        while current_date <= end_date:
            has_lesson = any(self._is_lesson_day(current_date, s) for s in schedule)

            if has_lesson:
                online_lesson = OnlineLesson.objects.filter(group=group, lesson_date=current_date).first()

                if online_lesson:
                    calendar_data.append({
                        'date':        current_date,
                        'has_topic':   True,
                        'topic':       online_lesson.title,
                        'description': online_lesson.description,
                        'has_content': self._check_content(online_lesson),
                        'lesson_id':   str(online_lesson.id),
                        'is_published': online_lesson.is_published,
                    })
                else:
                    calendar_data.append({
                        'date':        current_date,
                        'has_topic':   False,
                        'topic':       None,
                        'description': None,
                        'has_content': False,
                        'lesson_id':   None,
                        'is_published': False,
                    })

            current_date += timedelta(days=1)

        return Response({
            'group':        {'id': str(group.id), 'name': group.name},
            'start_date':   start_date,
            'end_date':     end_date,
            'total_lessons': len([d for d in calendar_data if d['has_topic']]),
            'calendar':     calendar_data,
        })

    def _is_lesson_day(self, date, schedule):
        weekday = date.weekday()
        if schedule.day_type == 'every':
            return True
        elif schedule.day_type == 'odd':
            return weekday in [0, 2, 4]
        elif schedule.day_type == 'even':
            return weekday in [1, 3, 5]
        return False

    def _check_content(self, lesson):
        if lesson.content_type == 'video':
            return bool(lesson.video_url)
        elif lesson.content_type in ['document', 'image']:
            return bool(lesson.file)
        elif lesson.content_type == 'link':
            return bool(lesson.external_link)
        elif lesson.content_type == 'text':
            return bool(lesson.text_content)
        return False


class BulkSetLessonTopicsView(APIView):
    """Ko'p darsga mavzu belgilash"""

    def post(self, request):
        serializer = BulkLessonTopicSerializer(data=request.data, context={'request': request})

        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        lessons = serializer.save()

        # ── AuditLog ──
        _log('other', None, 'create', None, {
            'action': f'{len(lessons)} ta dars mavzusi belgilandi (bulk)',
        }, request.user)

        return Response({
            'message': f'{len(lessons)} ta dars yaratildi',
            'count':   len(lessons),
            'lessons': OnlineLessonListSerializer(lessons, many=True).data,
        }, status=status.HTTP_201_CREATED)


class PublishLessonView(APIView):
    """Darsni nashr qilish/yashirish"""

    def post(self, request, lesson_id):
        lesson = get_object_or_404(OnlineLesson, id=lesson_id)

        has_content = False
        if lesson.content_type == 'video':
            has_content = bool(lesson.video_url)
        elif lesson.content_type in ['document', 'image']:
            has_content = bool(lesson.file)
        elif lesson.content_type == 'link':
            has_content = bool(lesson.external_link)
        elif lesson.content_type == 'text':
            has_content = bool(lesson.text_content)

        if not has_content:
            return Response({"error": "Darsda kontent yo'q. Avval kontent qo'shing."},
                            status=status.HTTP_400_BAD_REQUEST)

        lesson.is_published = True
        lesson.save()

        return Response({'message': 'Dars nashr qilindi', 'lesson': OnlineLessonDetailSerializer(lesson).data})

    def delete(self, request, lesson_id):
        lesson             = get_object_or_404(OnlineLesson, id=lesson_id)
        lesson.is_published = False
        lesson.save()
        return Response({'message': 'Dars yashirildi'})


class LessonStatisticsView(APIView):
    """Darslar statistikasi"""

    def get(self, request, group_id):
        group   = get_object_or_404(Group, id=group_id)
        lessons = OnlineLesson.objects.filter(group=group)

        total_lessons       = lessons.count()
        published_lessons   = lessons.filter(is_published=True).count()
        unpublished_lessons = total_lessons - published_lessons

        content_stats = {}
        for content_type, label in OnlineLesson.CONTENT_TYPE:
            content_stats[content_type] = {
                'label': label,
                'count': lessons.filter(content_type=content_type).count(),
            }

        return Response({
            'group':               {'id': str(group.id), 'name': group.name},
            'total_lessons':       total_lessons,
            'published_lessons':   published_lessons,
            'unpublished_lessons': unpublished_lessons,
            'content_by_type':     content_stats,
        })


class OnlineLessonCreateView(generics.CreateAPIView):
    serializer_class = OnlineLessonCreateUpdateSerializer

    def get_queryset(self):
        group_id  = self.request.query_params.get('group')
        course_id = self.request.query_params.get('course')
        if group_id:
            return OnlineLesson.objects.filter(group_id=group_id)
        if course_id:
            return OnlineLesson.objects.filter(group__course_id=course_id)
        return OnlineLesson.objects.all()

    def create(self, request, *args, **kwargs):
        course_id = request.data.get('course')
        if course_id:
            groups          = Group.objects.filter(course_id=course_id)
            created_lessons = []
            for group in groups:
                data          = request.data.copy()
                data['group'] = group.id
                serializer    = self.get_serializer(data=data)
                serializer.is_valid(raise_exception=True)
                serializer.save()
                created_lessons.append(serializer.data)
            return Response(created_lessons, status=status.HTTP_201_CREATED)
        return super().create(request, *args, **kwargs)


class OnlineLessonDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset         = OnlineLesson.objects.all()
    serializer_class = OnlineLessonDetailSerializer


# ════════════════════════════════════════════════════════════════
#  TALABALAR API
# ════════════════════════════════════════════════════════════════

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def student_list_create(request):
    if request.method == 'GET':
        qs = Student.objects.prefetch_related(
            'studentgroup_student__group__course',
            'student_balances',
            'studentfreezes_student',
        ).all()

        search = request.query_params.get('search')
        if search:
            qs = qs.filter(Q(full_name__icontains=search) | Q(phone_number__icontains=search))

        course_id = request.query_params.get('course')
        if course_id:
            ids = StudentGroup.objects.filter(
                group__course_id=course_id, left_at__isnull=True
            ).values_list('student_id', flat=True)
            qs = qs.filter(id__in=ids)

        student_id = request.query_params.get('id')
        if student_id:
            qs = qs.filter(id=student_id)

        date_from = request.query_params.get('date_from')
        date_to   = request.query_params.get('date_to')
        if date_from:
            qs = qs.filter(created_at__date__gte=date_from)
        if date_to:
            qs = qs.filter(created_at__date__lte=date_to)

        financial = request.query_params.get('financial')
        if financial == 'debtor':
            ids = StudentBalances.objects.filter(balance__lt=0).values_list('student_id', flat=True)
            qs  = qs.filter(id__in=ids)
        elif financial == 'positive':
            ids = StudentBalances.objects.filter(balance__gt=0).values_list('student_id', flat=True)
            qs  = qs.filter(id__in=ids)
        elif financial == 'zero':
            ids = StudentBalances.objects.filter(balance=0).values_list('student_id', flat=True)
            qs  = qs.filter(id__in=ids)

        student_status = request.query_params.get('status')
        if student_status == 'active':
            active_ids = StudentGroup.objects.filter(left_at__isnull=True).values_list('student_id', flat=True)
            frozen_ids = StudentFreezes.objects.filter(freeze_end_date__isnull=True).values_list('student_id', flat=True)
            qs = qs.filter(id__in=active_ids).exclude(id__in=frozen_ids)
        elif student_status == 'frozen':
            ids = StudentFreezes.objects.filter(freeze_end_date__isnull=True).values_list('student_id', flat=True)
            qs  = qs.filter(id__in=ids)
        elif student_status == 'left':
            active_ids = StudentGroup.objects.filter(left_at__isnull=True).values_list('student_id', flat=True)
            qs = qs.exclude(id__in=active_ids)
        elif student_status == 'trial':
            trial_date = timezone.now().date() - timedelta(days=3)
            ids = StudentGroup.objects.filter(
                left_at__isnull=True, joined_at__gte=trial_date
            ).values_list('student_id', flat=True)
            qs = qs.filter(id__in=ids)

        total  = qs.count()
        limit  = int(request.query_params.get('limit', 20))
        page   = int(request.query_params.get('page', 1))
        offset = (page - 1) * limit
        qs     = qs[offset: offset + limit]

        return Response({
            'count':   total,
            'page':    page,
            'limit':   limit,
            'results': StudentListSerializer(qs, many=True).data,
        })

    # ── POST ──
    serializer = StudentCreateSerializer(data=request.data)
    if serializer.is_valid():
        student = serializer.save()
        StudentBalances.objects.create(student=student, balance=0)

        # ── AuditLog ──
        _log('student', student.id, 'create', None, {
            'full_name':    student.full_name,
            'phone_number': student.phone_number,
        }, request.user)

        return Response(StudentDetailSerializer(student).data, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET', 'PUT', 'PATCH', 'DELETE'])
@permission_classes([IsAuthenticated])
def student_detail_update_delete(request, pk):
    student = get_object_or_404(Student, pk=pk)

    if request.method == 'GET':
        return Response(StudentDetailSerializer(student).data)

    elif request.method in ['PUT', 'PATCH']:
        old_data   = {'full_name': student.full_name, 'phone_number': student.phone_number}
        partial    = (request.method == 'PATCH')
        serializer = StudentUpdateSerializer(student, data=request.data, partial=partial)
        if serializer.is_valid():
            serializer.save()

            # ── AuditLog ──
            _log('student', student.id, 'update', old_data, {
                'full_name':    student.full_name,
                'phone_number': student.phone_number,
            }, request.user)

            return Response(StudentDetailSerializer(student).data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    elif request.method == 'DELETE':
        # ── AuditLog ──
        _log('student', student.id, 'delete',
             {'full_name': student.full_name, 'phone_number': student.phone_number},
             None, request.user)

        student.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


# ── Tablar ──────────────────────────────────────────────────────

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def student_groups_tab(request, pk):
    student = get_object_or_404(Student, pk=pk)
    sgs     = student.studentgroup_student.select_related('group__course').all()
    return Response(StudentGroupDetailSerializer(sgs, many=True).data)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def student_transactions_tab(request, pk):
    student  = get_object_or_404(Student, pk=pk)
    txns     = student.student_transactions.order_by('-transaction_date')
    txn_type = request.query_params.get('type')
    if txn_type:
        txns = txns.filter(transaction_type=txn_type)
    return Response({'count': txns.count(), 'results': TransactionSerializer(txns, many=True).data})


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def student_sms_tab(request, pk):
    student = get_object_or_404(Student, pk=pk)
    try:
        from communication.models import SMSMessages
        sms_list = SMSMessages.objects.filter(recipent_id=student.id).order_by('-created_at')
        data = [{'id': str(s.id), 'text': s.text, 'status': s.status,
                 'sent_at': s.sent_at, 'send_type': s.send_type} for s in sms_list]
    except Exception:
        data = []
    return Response({'count': len(data), 'results': data})


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def student_calls_tab(request, pk):
    student = get_object_or_404(Student, pk=pk)
    try:
        from finance.models import CallLog
        from finance.serializers import CallLogSerializer
        calls = CallLog.objects.filter(receiver_phone=student.phone_number).order_by('-call_time')
        return Response({'count': calls.count(),
                         'results': CallLogSerializer(calls, many=True, context={'request': request}).data})
    except Exception:
        return Response({'count': 0, 'results': []})


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def student_history_tab(request, pk):
    student = get_object_or_404(Student, pk=pk)
    try:
        logs = AuditLog.objects.filter(entity_id=student.id).order_by('-created_at')[:50]
        data = [{'id': str(l.id), 'action': l.action, 'entity_type': l.entity_type,
                 'old_data': l.old_data, 'new_data': l.new_action,
                 'performed_by': l.performed_by.user.full_name if l.performed_by else None,
                 'created_at': l.created_at} for l in logs]
    except Exception:
        data = []
    return Response({'count': len(data), 'results': data})


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def student_lead_history_tab(request, pk):
    student = get_object_or_404(Student, pk=pk)
    try:
        from crm.models import CRMLead
        lead = CRMLead.objects.filter(converted_student_id=student.id).first()
        if not lead:
            return Response({'lead': None, 'activities': []})
        return Response({
            'lead': {'id': str(lead.id), 'full_name': lead.full_name,
                     'phone_number': lead.phone_number, 'status': lead.status,
                     'source': lead.source.name if lead.source else None,
                     'pipeline': lead.pipline.name if lead.pipline else None},
            'activities': [{'id': str(a.id), 'activity_type': a.activity_type,
                             'result': a.result, 'created_at': a.created_at}
                           for a in lead.activities.order_by('-created_at')],
        })
    except Exception:
        return Response({'lead': None, 'activities': []})


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def student_coin_history_tab(request, pk):
    student = get_object_or_404(Student, pk=pk)
    history = StudentBalanceHistory.objects.filter(student=student).order_by('-created_at')[:50]
    data = [{'id': str(h.id), 'amount': float(h.amount),
             'base_price': float(h.base_price) if h.base_price else None,
             'applied_price': float(h.applied_price) if h.applied_price else None,
             'discount': float(h.discount) if h.discount else None,
             'created_at': h.created_at} for h in history]
    return Response({'count': len(data), 'results': data})

# ════════════════════════════════════════════════════════════════
#  AMALLAR  (function-based)
# ════════════════════════════════════════════════════════════════

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def student_add_payment(request, pk):
    """Body: { amount, payment_type, comment, group_id (optional) }"""
    student      = get_object_or_404(Student, pk=pk)
    amount       = request.data.get('amount')
    payment_type = request.data.get('payment_type', 'cash')
    comment      = request.data.get('comment', '')
    group_id     = request.data.get('group_id')

    if not amount:
        return Response({'error': 'amount majburiy'}, status=status.HTTP_400_BAD_REQUEST)

    group = get_object_or_404(Group, pk=group_id) if group_id else None

    txn = StudentTarnsactions.objects.create(
        student=student, group=group,
        transaction_type='payment', amount=amount,
        payment_type=payment_type,
        transaction_date=timezone.now(),
        comment=comment,
        accepted_by=request.user.employee,
    )

    balance_obj, _ = StudentBalances.objects.get_or_create(student=student, defaults={'balance': 0})
    old_balance     = float(balance_obj.balance)
    balance_obj.balance += float(amount)
    balance_obj.save()

    # ── AuditLog ──
    _log('payment', student.id, 'create', {'balance': old_balance}, {
        'amount':         float(amount),
        'payment_type':   payment_type,
        'new_balance':    float(balance_obj.balance),
        'comment':        comment,
        'group_id':       str(group_id) if group_id else None,
        'transaction_id': str(txn.id),
    }, request.user)

    return Response({
        'success':        True,
        'message':        f'+{amount} UZS to\'lov qo\'shildi',
        'new_balance':    float(balance_obj.balance),
        'transaction_id': str(txn.id),
    }, status=status.HTTP_201_CREATED)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def student_add_to_group(request, pk):
    """Body: { group_id }"""
    student  = get_object_or_404(Student, pk=pk)
    group_id = request.data.get('group_id')
    if not group_id:
        return Response({'error': 'group_id majburiy'}, status=400)

    group = get_object_or_404(Group, pk=group_id)
    if StudentGroup.objects.filter(student=student, group=group, left_at__isnull=True).exists():
        return Response({'error': 'Talaba bu guruhda allaqachon mavjud'}, status=400)

    sg = StudentGroup.objects.create(student=student, group=group, joined_at=timezone.now().date())

    # ── AuditLog ──
    _log('student', student.id, 'update', None, {
        'action':    "Guruhga qo'shildi",
        'group_id':  str(group.id),
        'group':     group.name,
        'joined_at': str(timezone.now().date()),
    }, request.user)

    return Response({
        'success':          True,
        'message':          f'Talaba "{group.name}" guruhiga qo\'shildi',
        'student_group_id': str(sg.id),
    }, status=status.HTTP_201_CREATED)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def student_remove_from_group(request, pk):
    """Body: { group_id, leave_reason_id (optional), comment (optional) }"""
    student  = get_object_or_404(Student, pk=pk)
    group_id = request.data.get('group_id')
    if not group_id:
        return Response({'error': 'group_id majburiy'}, status=400)

    sg = StudentGroup.objects.filter(student=student, group_id=group_id, left_at__isnull=True).first()
    if not sg:
        return Response({'error': 'Talaba bu guruhda topilmadi'}, status=404)

    sg.left_at = timezone.now().date()
    sg.save()

    leave_reason_id = request.data.get('leave_reason_id')
    comment         = request.data.get('comment', '')
    try:
        leave_reason = LeaveReason.objects.filter(pk=leave_reason_id).first() if leave_reason_id else None
        StudentGroupLeaves.objects.create(
            student=student, group=sg.group, student_group=sg,
            leave_date=timezone.now(), leave_reason=leave_reason,
            comment=comment, created_by=request.user.employee,
        )
    except Exception:
        pass

    # ── AuditLog ──
    _log('student', student.id, 'update',
         {'group_id': str(group_id)},
         {
             'action':          "Guruhdan chiqarildi",
             'group_id':        str(group_id),
             'group':           sg.group.name,
             'leave_date':      str(timezone.now().date()),
             'leave_reason_id': str(leave_reason_id) if leave_reason_id else None,
             'comment':         comment,
         }, request.user)

    return Response({'success': True, 'message': f'Talaba "{sg.group.name}" guruhidan chiqarildi'})


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def student_freeze(request, pk):
    """Body: { group_id, freeze_start_date, freeze_end_date, reason }"""
    student           = get_object_or_404(Student, pk=pk)
    group_id          = request.data.get('group_id')
    freeze_start_date = request.data.get('freeze_start_date')
    freeze_end_date   = request.data.get('freeze_end_date')
    reason            = request.data.get('reason', '')

    if not group_id or not freeze_start_date or not freeze_end_date:
        return Response({'error': 'group_id, freeze_start_date, freeze_end_date majburiy'}, status=400)

    group = get_object_or_404(Group, pk=group_id)

    if StudentFreezes.objects.filter(student=student, group=group, freeze_end_date__isnull=True).exists():
        return Response({'error': 'Talaba allaqachon muzlatilgan'}, status=400)

    freeze = StudentFreezes.objects.create(
        student=student, group=group,
        freeze_start_date=freeze_start_date,
        freeze_end_date=freeze_end_date,
        reason=reason,
        created_by=request.user.employee,
    )

    # ── AuditLog ──
    _log('student', student.id, 'update', None, {
        'action':            'Muzlatildi',
        'group_id':          str(group.id),
        'group':             group.name,
        'freeze_start_date': str(freeze_start_date),
        'freeze_end_date':   str(freeze_end_date),
        'reason':            reason,
        'freeze_id':         str(freeze.id),
    }, request.user)

    return Response({
        'success':    True,
        'message':    'Talaba muzlatildi',
        'freeze_id':  str(freeze.id),
        'start_date': freeze_start_date,
        'end_date':   freeze_end_date,
    }, status=status.HTTP_201_CREATED)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def student_unfreeze(request, pk):
    """Body: { group_id }"""
    student  = get_object_or_404(Student, pk=pk)
    group_id = request.data.get('group_id')

    freeze = StudentFreezes.objects.filter(
        student=student, group_id=group_id, freeze_end_date__isnull=True
    ).first()
    if not freeze:
        return Response({'error': 'Faol muzlatish topilmadi'}, status=404)

    old_end_date           = freeze.freeze_end_date
    freeze.freeze_end_date = timezone.now().date()
    freeze.save()

    # ── AuditLog ──
    _log('student', student.id, 'update',
         {'freeze_end_date': str(old_end_date)},
         {
             'action':          'Muzlatish bekor qilindi',
             'group_id':        str(group_id),
             'freeze_end_date': str(freeze.freeze_end_date),
         }, request.user)

    return Response({'success': True, 'message': 'Muzlatish bekor qilindi'})


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def student_send_sms(request, pk):
    """Body: { text, template_id (optional) }"""
    student = get_object_or_404(Student, pk=pk)
    text    = request.data.get('text')

    if not text:
        return Response({'error': 'text majburiy'}, status=400)
    if not student.phone_number:
        return Response({'error': "Talabaning telefon raqami yo'q"}, status=400)

    try:
        from communication.models import SMSMessages, SmsTemplates
        template = SmsTemplates.objects.filter(pk=request.data.get('template_id')).first()
        sms = SMSMessages.objects.create(
            recipent_type='student', recipent_id=student.id,
            phone=student.phone_number, text=text,
            template=template, send_type='manual',
            status='pending', sent_at=timezone.now(),
        )
        return Response({
            'success': True,
            'message': f'SMS yuborildi: {student.phone_number}',
            'sms_id':  str(sms.id),
        }, status=status.HTTP_201_CREATED)
    except Exception as e:
        return Response({'error': str(e)}, status=500)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def student_export_excel(request):
    try:
        from openpyxl.styles import Font, PatternFill, Alignment

        qs = Student.objects.prefetch_related('studentgroup_student__group', 'student_balances').all()
        search    = request.query_params.get('search')
        course_id = request.query_params.get('course')
        fin       = request.query_params.get('financial')

        if search:
            qs = qs.filter(Q(full_name__icontains=search) | Q(phone_number__icontains=search))
        if course_id:
            ids = StudentGroup.objects.filter(group__course_id=course_id, left_at__isnull=True).values_list('student_id', flat=True)
            qs  = qs.filter(id__in=ids)
        if fin == 'debtor':
            ids = StudentBalances.objects.filter(balance__lt=0).values_list('student_id', flat=True)
            qs  = qs.filter(id__in=ids)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Talabalar'

        headers     = ['#', 'ID', 'Ism', 'Telefon', 'Guruh', 'Balans', 'Coins', "Qo'shilgan sana"]
        header_fill = PatternFill(start_color='FF6B35', end_color='FF6B35', fill_type='solid')
        for col, h in enumerate(headers, 1):
            cell           = ws.cell(row=1, column=col, value=h)
            cell.font      = Font(bold=True, color='FFFFFF')
            cell.fill      = header_fill
            cell.alignment = Alignment(horizontal='center')

        for row_num, student in enumerate(qs, 2):
            b          = student.student_balances.first()
            sg         = student.studentgroup_student.filter(left_at__isnull=True).first()
            ws.cell(row=row_num, column=1, value=row_num - 1)
            ws.cell(row=row_num, column=2, value=student.id)
            ws.cell(row=row_num, column=3, value=student.full_name)
            ws.cell(row=row_num, column=4, value=student.phone_number)
            ws.cell(row=row_num, column=5, value=sg.group.name if sg else '-')
            ws.cell(row=row_num, column=6, value=float(b.balance) if b else 0)
            ws.cell(row=row_num, column=7, value=student.coins)
            ws.cell(row=row_num, column=8, value=str(student.created_at.date()))

        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = max(len(str(c.value or '')) for c in col) + 4

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="students.xlsx"'
        wb.save(response)
        return response

    except ImportError:
        return Response({'error': 'openpyxl o\'rnatilmagan: pip install openpyxl'}, status=500)
    except Exception as e:
        return Response({'error': str(e)}, status=500)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def student_filter_options(request):
    return Response({
        'courses': list(Course.objects.values('id', 'name')),
        'statuses': [
            {'value': 'active',  'label': 'Faol'},
            {'value': 'trial',   'label': 'Sinov darsida'},
            {'value': 'frozen',  'label': 'Muzlatilgan'},
            {'value': 'left',    'label': 'Chiqib ketgan'},
        ],
        'financial_statuses': [
            {'value': 'debtor',   'label': 'Qarzdor (manfiy)'},
            {'value': 'zero',     'label': 'Nol balansi'},
            {'value': 'positive', 'label': 'Musbat balans'},
        ],
    })


# ════════════════════════════════════════════════════════════════
#  CLASS-BASED VIEWS (qo'shimcha)
# ════════════════════════════════════════════════════════════════

class GroupListView(APIView):
    """Guruhlar ro'yxati — filterlar bilan"""

    def get(self, request):
        qs = Group.objects.select_related('course', 'room').all()

        if s := request.GET.get('status'):
            qs = qs.filter(status=s)
        if c := request.GET.get('course'):
            qs = qs.filter(course_id=c)
        if q := request.GET.get('search'):
            qs = qs.filter(name__icontains=q)

        from .serializers import GroupListSerializer
        return Response({'success': True, 'count': qs.count(),
                         'results': GroupListSerializer(qs[:50], many=True).data})


class GroupStudentsView(APIView):
    """Guruhdagi talabalar ro'yxati"""

    def get(self, request, group_id):
        group         = get_object_or_404(Group, pk=group_id)
        status_filter = request.GET.get('status', 'active')
        qs            = StudentGroup.objects.filter(group=group)

        if status_filter == 'active':
            qs = qs.filter(left_at__isnull=True)
        elif status_filter == 'left':
            qs = qs.filter(left_at__isnull=False)

        from .serializers import StudentGroupSerializer
        return Response(StudentGroupSerializer(qs, many=True).data)


class StudentSearchView(APIView):
    """Talaba qidirish"""

    def get(self, request):
        from .serializers import StudentSearchSerializer
        q = request.GET.get('q', '').strip()
        if not q:
            return Response({'error': "q parametri bo'sh"}, status=400)

        students = Student.objects.filter(
            Q(full_name__icontains=q) | Q(phone_number__icontains=q)
        )[:15]

        return Response({'success': True, 'count': students.count(),
                         'results': StudentSearchSerializer(students, many=True).data})


class StudentAddPaymentView(APIView):
    """Talabaga to'lov qo'shish"""

    def post(self, request, student_id):
        student = get_object_or_404(Student, pk=student_id)
        amount  = request.data.get('amount')

        if not amount or float(amount) <= 0:
            return Response({'error': "To'g'ri summa kiriting"}, status=400)

        balance, _  = StudentBalances.objects.get_or_create(student=student, defaults={'balance': 0})
        old_balance = float(balance.balance)
        balance.balance += float(amount)
        balance.save()

        StudentTarnsactions.objects.create(
            student=student, amount=amount,
            transaction_type='payment',
            payment_type=request.data.get('payment_type', 'cash'),
            comment=request.data.get('comment', ''),
            accepted_by=request.user.employee if hasattr(request.user, 'employee') else None,
        )

        # ── AuditLog ──
        _log('payment', student.id, 'create', {'balance': old_balance}, {
            'amount':      float(amount),
            'new_balance': float(balance.balance),
        }, request.user)

        return Response({'success': True, 'new_balance': float(balance.balance),
                         'message': f"+{amount} so'm qo'shildi"})


class StudentDetailView(APIView):
    """Talabaning to'liq kartasi"""

    def get(self, request, pk):
        student     = get_object_or_404(Student, pk=pk)
        balance_obj = student.student_balances.first()
        balance     = float(balance_obj.balance) if balance_obj else 0

        active_groups = StudentGroup.objects.filter(student=student, left_at__isnull=True).select_related('group__course')
        transactions  = student.student_transactions.order_by('-transaction_date')[:10]
        discounts     = StudentPricing.objects.filter(student=student, end_date__gte=timezone.now().date())

        return Response({'success': True, 'data': {
            'id':           str(student.id),
            'full_name':    student.full_name,
            'phone_number': student.phone_number,
            'balance':      balance,
            'active_groups': [{'group_id': str(g.group.id), 'group_name': g.group.name,
                                'course': g.group.course.name if g.group.course else None,
                                'joined_at': g.joined_at} for g in active_groups],
            'recent_transactions': [{'amount': float(t.amount), 'type': t.transaction_type,
                                     'date': t.transaction_date, 'comment': t.comment} for t in transactions],
            'active_discounts': [{'course': d.course.name, 'price_override': float(d.price_override),
                                  'reason': d.reason} for d in discounts],
            'status': 'active' if balance > 0 else 'inactive',
        }})


class GroupAttendanceView(APIView):
    """Guruhning tanlangan kungi davomati"""

    def get(self, request, group_id):
        group           = get_object_or_404(Group, pk=group_id)
        lesson_date_str = request.GET.get('date')
        try:
            lesson_date = date.fromisoformat(lesson_date_str) if lesson_date_str else date.today()
        except ValueError:
            return Response({'error': "Noto'g'ri sana formati (YYYY-MM-DD)"}, status=400)

        student_groups = StudentGroup.objects.filter(group=group, left_at__isnull=True).select_related('student')
        attendances    = Attendence.objects.filter(student_group__in=student_groups, lesson_date=lesson_date)

        data = []
        for sg in student_groups:
            att = attendances.filter(student_group=sg).first()
            data.append({
                'student_group_id': str(sg.id),
                'student_name':     sg.student.full_name,
                'phone':            sg.student.phone_number,
                'is_present':       att.is_present if att else None,
                'marked_at':        att.created_at if att else None,
            })

        return Response({'success': True, 'date': str(lesson_date), 'total_students': len(data), 'results': data})


class AttendanceSaveView(APIView):
    """Bir kunlik davomatni saqlash"""

    def post(self, request, group_id):
        group           = get_object_or_404(Group, pk=group_id)
        lesson_date_str = request.data.get('date')
        if not lesson_date_str:
            return Response({'error': 'date parametri majburiy'}, status=400)

        try:
            lesson_date = date.fromisoformat(lesson_date_str)
        except ValueError:
            return Response({'error': "Sana formati noto'g'ri (YYYY-MM-DD)"}, status=400)

        attendances_data = request.data.get('attendances', [])
        if not attendances_data:
            return Response({'error': "attendances ro'yxati bo'sh"}, status=400)

        saved = 0
        for item in attendances_data:
            try:
                sg = StudentGroup.objects.get(id=item.get('student_group'), group=group)
                Attendence.objects.update_or_create(
                    student_group=sg, lesson_date=lesson_date,
                    defaults={
                        'is_present': item.get('is_present', False),
                        'marked_by':  request.user.employee if hasattr(request.user, 'employee') else None,
                    }
                )
                saved += 1
            except StudentGroup.DoesNotExist:
                pass

        return Response({'success': True, 'saved_count': saved,
                         'message': f'{saved} ta talabaning davomati saqlandi'}, status=201)


class GroupDiscountsView(APIView):
    """Guruhdagi talabalarga qo'yilgan chegirmalar"""

    def get(self, request, group_id):
        group     = get_object_or_404(Group, pk=group_id)
        discounts = StudentPricing.objects.filter(
            student__studentgroup_student__group=group,
            student__studentgroup_student__left_at__isnull=True,
        ).select_related('student', 'course')

        data = [{'student': d.student.full_name, 'student_id': str(d.student.id),
                 'original_price': float(d.course.monthly_price),
                 'discount_price': float(d.price_override),
                 'start_date': d.start_date, 'end_date': d.end_date, 'reason': d.reason}
                for d in discounts]

        return Response({'success': True, 'count': len(data), 'results': data})


class SetStudentDiscountView(APIView):
    """Talabaga chegirma / maxsus narx qo'yish"""

    def post(self, request, group_id, sg_id):
        group   = get_object_or_404(Group, pk=group_id)
        sg      = get_object_or_404(StudentGroup, pk=sg_id, group=group)
        student = sg.student
        course  = group.course
        price   = request.data.get('price_override')

        if not price:
            return Response({'error': 'price_override majburiy'}, status=400)

        today  = date.today()
        start  = request.data.get('start_date', today.isoformat())
        end    = request.data.get('end_date', group.end_date.isoformat() if group.end_date else None)
        reason = request.data.get('reason', '')

        old_prices = []
        for o in StudentPricing.objects.filter(student=student, course=course, end_date__gte=today):
            old_prices.append(str(o.price_override))
            o.end_date = today
            o.save()

        discount = StudentPricing.objects.create(
            student=student, course=course,
            price_override=price, reason=reason,
            start_date=start, end_date=end,
            created_by=request.user.employee if hasattr(request.user, 'employee') else None,
        )

        # ── AuditLog ──
        _log('student', student.id, 'update',
             {'old_prices': old_prices},
             {'action': 'Chegirma qo\'yildi (SetStudentDiscountView)',
              'price_override': str(discount.price_override),
              'course_id': str(course.id)},
             request.user)

        return Response({'success': True, 'message': f"{student.full_name} uchun chegirma qo'yildi",
                         'discount_price': float(discount.price_override)}, status=201)


class ExamListView(APIView):
    """Imtihonlar ro'yxati"""

    def get(self, request):
        qs       = Exams.objects.prefetch_related('group').all()
        group_id = request.GET.get('group')
        if group_id:
            qs = qs.filter(group_id=group_id)

        data = [{'id': str(e.id), 'title': e.title, 'group': e.group.name,
                 'exam_date': e.exam_date, 'min_score': float(e.min_score),
                 'max_score': float(e.max_score)} for e in qs]

        return Response({'success': True, 'count': len(data), 'results': data})


class ExamResultsSubmitView(APIView):
    """Imtihon natijalarini kiritish"""

    def post(self, request, exam_id):
        exam         = get_object_or_404(Exams, pk=exam_id)
        results_data = request.data.get('results', [])
        if not results_data:
            return Response({'error': "results ro'yxati bo'sh"}, status=400)

        saved = []
        for item in results_data:
            try:
                student = Student.objects.get(pk=item.get('student'))
                result, _ = ExamResults.objects.update_or_create(
                    exam=exam, student=student,
                    defaults={
                        'score':      item.get('score'),
                        'comment':    item.get('comment', ''),
                        'created_by': request.user.employee if hasattr(request.user, 'employee') else None,
                    }
                )
                saved.append({'student': student.full_name, 'score': float(item.get('score'))})
            except Student.DoesNotExist:
                pass

        # ── AuditLog ──
        if saved:
            _log('other', exam.id, 'update', None, {
                'action':      'Imtihon natijalari kiritildi (ExamResultsSubmitView)',
                'saved_count': len(saved),
            }, request.user)

        return Response({'success': True, 'saved': len(saved), 'results': saved}, status=201)


class GroupOnlineLessonsView(APIView):
    """Guruhning onlayn darslari"""

    def get(self, request, group_id):
        group   = get_object_or_404(Group, pk=group_id)
        lessons = OnlineLesson.objects.filter(group=group).order_by('order', '-lesson_date')

        data = [{'id': str(l.id), 'title': l.title, 'lesson_date': l.lesson_date,
                 'content_type': l.content_type, 'is_published': l.is_published,
                 'duration': l.duration_minutes} for l in lessons]

        return Response({'success': True, 'count': len(data), 'results': data})


class GroupUpdateView(APIView):
    """Guruhni tahrirlash"""

    def patch(self, request, group_id):
        from .serializers import GroupUpdateSerializer, GroupDetailSerializer
        group    = get_object_or_404(Group, pk=group_id)
        old_data = {'name': group.name, 'status': group.status}

        serializer = GroupUpdateSerializer(group, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()

            # ── AuditLog ──
            _log('group', group.id, 'update', old_data,
                 {'name': group.name, 'status': group.status}, request.user)

            return Response({'success': True, 'message': "Guruh ma'lumotlari yangilandi",
                             'group': GroupDetailSerializer(group).data})
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class GroupSendSMSView(APIView):
    """Guruhdagi talabalarga SMS yuborish"""

    def post(self, request, group_id):
        group   = get_object_or_404(Group, pk=group_id)
        message = request.data.get('message')
        send_to = request.data.get('send_to', 'all')

        if not message:
            return Response({'error': 'Xabar matni majburiy'}, status=400)

        phones   = []
        students = StudentGroup.objects.filter(group=group, left_at__isnull=True)
        for sg in students:
            s = sg.student
            if send_to in ['all', 'students'] and s.phone_number:
                phones.append(s.phone_number)
            if send_to in ['all', 'parents'] and s.parent_phone:
                phones.append(s.parent_phone)

        # TODO: real SMS provayder
        return Response({'success': True, 'message': f'{len(phones)} ta raqamga SMS yuborildi',
                         'phones_count': len(phones)})


class GroupAddNewStudentView(APIView):
    """Yangi talaba yaratish va guruhga qo'shish"""

    def post(self, request, group_id):
        from .serializers import StudentCreateSerializer, StudentDetailSerializer
        group      = get_object_or_404(Group, pk=group_id)
        serializer = StudentCreateSerializer(data=request.data)

        if not serializer.is_valid():
            return Response(serializer.errors, status=400)

        student   = serializer.save()
        joined_at = request.data.get('joined_at', timezone.now().date())
        sg        = StudentGroup.objects.create(student=student, group=group, joined_at=joined_at)
        StudentBalances.objects.get_or_create(student=student, defaults={'balance': 0})

        # ── AuditLog ──
        _log('student', student.id, 'create', None, {
            'action':    "Yangi talaba yaratildi va guruhga qo'shildi",
            'full_name': student.full_name,
            'group_id':  str(group.id),
            'group':     group.name,
        }, request.user)

        return Response({
            'success':          True,
            'message':          "Yangi talaba yaratildi va guruhga qo'shildi",
            'student':          StudentDetailSerializer(student).data,
            'student_group_id': str(sg.id),
        }, status=201)